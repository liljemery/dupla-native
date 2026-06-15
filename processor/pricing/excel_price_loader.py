"""
Loader for the constructor pricing Excel (formato Dupla Constructora).

Detects the three expected sheets and parses each into the typed dataclasses
defined in :mod:`pricing.schemas`.

Sheet routing:
    * "analisis ..."           -> APUs       (APUBreakdown)
    * "lista ..." / "precios " -> Materiales (MaterialPrice)
    * "MO ..."   / "mano ..."  -> M. de obra (LaborRate)
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any, Iterator

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .schemas import (
    APUBreakdown,
    APUComponent,
    LaborRate,
    MaterialPrice,
    PricingStore,
    PricingExcelConfig,
)

logger = logging.getLogger("dupla.pricing.excel_loader")


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

DEFAULT_CACHE_DIR = Path(__file__).resolve().parent.parent / "data" / "pricing_cache"


def save_pricing_store(store: PricingStore, json_path: str | Path) -> Path:
    """Serialize a :class:`PricingStore` to a JSON cache file."""
    path = Path(json_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(store.to_dict(), ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    return path


def load_pricing_store(json_path: str | Path) -> PricingStore:
    """Load a cached :class:`PricingStore` from JSON."""
    path = Path(json_path)
    payload = json.loads(path.read_text(encoding="utf-8"))
    return PricingStore.from_dict(payload)


def cache_path_for(project_id: str, cache_dir: str | Path | None = None) -> Path:
    base = Path(cache_dir) if cache_dir else DEFAULT_CACHE_DIR
    return base / f"{project_id}_pricing.json"


def load_or_cache_constructor_pricing(
    excel_path: str | Path,
    project_id: str,
    *,
    cache_dir: str | Path | None = None,
    force_refresh: bool = False,
    config: PricingExcelConfig | None = None,
) -> PricingStore:
    """
    Return a :class:`PricingStore` for ``project_id``, using a JSON cache when
    the Excel has not changed since the cache was written.

    The cache file is invalidated by either an older mtime than the Excel
    workbook, or by an explicit ``force_refresh``.
    """
    excel = Path(excel_path)
    cache_file = cache_path_for(project_id, cache_dir)

    if not force_refresh and cache_file.exists() and excel.exists():
        if cache_file.stat().st_mtime >= excel.stat().st_mtime:
            try:
                store = load_pricing_store(cache_file)
                logger.info("PricingStore loaded from cache: %s", cache_file)
                return store
            except Exception:
                logger.warning("Cache read failed, reparsing Excel", exc_info=True)

    store = load_constructor_pricing(str(excel), config=config)
    store.metadata["project_id"] = project_id
    save_pricing_store(store, cache_file)
    logger.info("PricingStore cached at: %s", cache_file)
    return store


def load_constructor_pricing(excel_path: str, config: PricingExcelConfig | None = None) -> PricingStore:
    """Parse the constructor pricing Excel into a :class:`PricingStore`."""
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Pricing Excel not found: {path}")

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    source = path.name

    if config is None:
        config = PricingExcelConfig()

    apus_sheet = _find_sheet(wb.sheetnames, config.apus.sheet_names)
    materials_sheet = _find_sheet(wb.sheetnames, config.materials.sheet_names)
    labor_sheet = _find_sheet(wb.sheetnames, config.labor.sheet_names)

    materials: dict[str, MaterialPrice] = {}
    labor: dict[str, LaborRate] = {}
    apus: dict[str, APUBreakdown] = {}

    if materials_sheet:
        materials = _parse_materials(wb[materials_sheet], source, config.materials)
        logger.info("Materials parsed: %d (sheet=%s)", len(materials), materials_sheet)
    else:
        logger.warning("Materials sheet not detected in %s", path)

    if labor_sheet:
        labor = _parse_labor(wb[labor_sheet], source, config.labor)
        logger.info("Labor activities parsed: %d (sheet=%s)", len(labor), labor_sheet)
    else:
        logger.warning("Labor sheet not detected in %s", path)

    if apus_sheet:
        apus = _parse_apus(wb[apus_sheet], source, config.apus)
        logger.info("APUs parsed: %d (sheet=%s)", len(apus), apus_sheet)
    else:
        logger.warning("APUs sheet not detected in %s", path)

    metadata: dict[str, Any] = {
        "source_file": str(path),
        "source_name": source,
        "sheet_apus": apus_sheet,
        "sheet_materials": materials_sheet,
        "sheet_labor": labor_sheet,
    }

    return PricingStore(materials=materials, labor=labor, apus=apus, metadata=metadata)


# ---------------------------------------------------------------------------
# Sheet detection
# ---------------------------------------------------------------------------

def _find_sheet(names: list[str], needles: tuple[str, ...]) -> str | None:
    for name in names:
        lo = name.lower()
        for needle in needles:
            if needle in lo:
                return name
    return None


# ---------------------------------------------------------------------------
# Generic helpers
# ---------------------------------------------------------------------------

def _norm_code(value: Any) -> str | None:
    """Normalise a code cell: float -> rounded 2-decimal string, str -> stripped."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        try:
            r = round(float(value), 2)
        except (TypeError, ValueError):
            return None
        if r == int(r):
            return str(int(r))
        return f"{r:.2f}"
    text = str(value).strip()
    return text or None


def _to_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    # Tolerate ranges like "1500 - 1600" -> take first number.
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def _norm_unit(value: Any) -> str:
    text = _norm_text(value)
    if not text:
        return ""
    # Excel-mangled superscripts: M³/M² often arrive as "M\udcb3" / "M\udcb2".
    text = text.replace("³", "3").replace("²", "2")
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def _is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in row)


# ---------------------------------------------------------------------------
# Materials sheet
# ---------------------------------------------------------------------------

def _parse_materials(ws: Worksheet, source: str, config: Any) -> dict[str, MaterialPrice]:
    """
    Columns of interest based on config.
    """
    result: dict[str, MaterialPrice] = {}
    current_category = ""
    auto_idx = 0

    max_col_needed = max(config.col_code, config.col_desc, config.col_unit, config.col_price, config.col_date) + 1

    for row in _iter_rows(ws, max_col=max_col_needed):
        if _is_blank_row(row):
            continue
        
        row_padded = row + (None,) * max_col_needed
        code_raw = row_padded[config.col_code]
        desc_raw = row_padded[config.col_desc]
        unit_raw = row_padded[config.col_unit]
        price_raw = row_padded[config.col_price]
        date_raw = row_padded[config.col_date]
        description = _norm_text(desc_raw)
        if not description:
            continue

        # Integer N° + no unit + no price = category header ("ACEROS-MALLAS-ALAMBRES").
        is_integer_code = isinstance(code_raw, (int, float)) and float(code_raw) == int(code_raw)
        unit_text = _norm_unit(unit_raw)
        price = _to_float(price_raw)

        if is_integer_code and not unit_text and price is None:
            current_category = description.upper()
            continue
        if description.upper() == "LISTA DE PRECIOS":
            continue
        if price is None or not unit_text:
            continue

        code = _norm_code(code_raw)
        if not code:
            auto_idx += 1
            code = f"mat_auto_{auto_idx:04d}"
        key = code
        # Avoid clobbering: if duplicate, suffix.
        if key in result:
            auto_idx += 1
            key = f"{code}__{auto_idx}"

        updated = None
        if hasattr(date_raw, "isoformat"):
            try:
                updated = date_raw.date().isoformat() if hasattr(date_raw, "date") else date_raw.isoformat()
            except Exception:
                updated = None

        result[key] = MaterialPrice(
            code=code,
            description=description,
            unit=unit_text,
            unit_price=round(price, 4),
            category=current_category,
            updated_date=updated,
            source=source,
        )

    return result


# ---------------------------------------------------------------------------
# Labor sheet
# ---------------------------------------------------------------------------

# MO 25 layout (0-indexed):
#   0: code (float)         1: description
#   7: CANTIDAD              9: UND
#  11: PU ($)               13: COSTO / PRESUPUESTO
_LABOR_HEADER_RE = re.compile(r"^\d+\.\d+-")    # e.g. "2.0-EXCAVACION A MANO"


def _parse_labor(ws: Worksheet, source: str, config: Any) -> dict[str, LaborRate]:
    result: dict[str, LaborRate] = {}
    current_category = ""
    auto_idx = 0

    max_col_needed = max(config.col_code, config.col_desc, config.col_qty, config.col_unit, config.col_price, config.col_price_fallback) + 1

    for row in _iter_rows(ws, max_col=max_col_needed):
        if _is_blank_row(row):
            continue
        row_padded = row + (None,) * max_col_needed
        code_raw = row_padded[config.col_code]
        description = _norm_text(row_padded[config.col_desc]) if row_padded[config.col_desc] is not None else ""

        # Category banners: text in col A like "EXCAVACIONES" or "2.0-EXCAVACION A MANO".
        if isinstance(code_raw, str):
            text_a = _norm_text(code_raw)
            if not text_a:
                continue
            if text_a.endswith(":") or _LABOR_HEADER_RE.match(text_a) or text_a.isupper():
                current_category = re.sub(r"^\d+\.\d+-\s*", "", text_a).upper()
                continue
            continue

        # Activity rows have either a numeric code or no code at all (continuation rows).
        if not description:
            continue

        unit_text = _norm_unit(row_padded[config.col_unit])
        price = _to_float(row_padded[config.col_price])
        if price is None:
            price = _to_float(row_padded[config.col_price_fallback])
        if price is None or not unit_text:
            continue

        code = _norm_code(code_raw)
        if not code:
            auto_idx += 1
            code = f"mo_auto_{auto_idx:04d}"
        key = code
        if key in result:
            auto_idx += 1
            key = f"{code}__{auto_idx}"

        result[key] = LaborRate(
            code=code,
            description=description,
            unit=unit_text,
            unit_price=round(price, 4),
            category=current_category,
            source=source,
        )

    return result


# ---------------------------------------------------------------------------
# APUs sheet
# ---------------------------------------------------------------------------

# analisis may25 layout (0-indexed, after row 6 header):
#   0: No.        1: PARTIDA / componente
#   2: CANT.      3: UD
#   4: P/UNIT.    5: VALOR (subtotal)
#   6: TOTAL PARTIDA   7: "/UNIT" sufijo
_APU_VALID_COMPONENT_TYPES = {"material", "labor", "equipment", "overhead"}


def _classify_component(description: str) -> str:
    desc_lo = description.lower()
    labor_kw = (
        "ayudante", "maestro", "personal", "operario", "obrero", "jornal", "mano",
        "envarillado", "carpinter", "encofrado", "armado", "replanteo", "vaciado",
    )
    equip_kw = ("retroexcavadora", "equipo", "bomba", "vibrador", "ligado", "mezcladora")
    overhead_kw = ("desperdicio", "supervision", "supervisión", "imprevisto", "gastos", "%")
    if any(k in desc_lo for k in labor_kw):
        return "labor"
    if any(k in desc_lo for k in equip_kw):
        return "equipment"
    if any(k in desc_lo for k in overhead_kw):
        return "overhead"
    return "material"


def _parse_apus(ws: Worksheet, source: str, config: Any) -> dict[str, APUBreakdown]:
    result: dict[str, APUBreakdown] = {}
    current_category = ""
    auto_idx = 0

    current_apu: dict[str, Any] | None = None

    def _close_current(total_value: float | None, total_unit: str | None) -> None:
        nonlocal current_apu, auto_idx
        if current_apu is None:
            return
        components: list[APUComponent] = current_apu["components"]
        unit = total_unit or current_apu.get("unit") or ""
        if total_value is None:
            total_value = round(sum(c.subtotal for c in components), 4) if components else 0.0
        code = current_apu["code"] or ""
        if not code:
            auto_idx += 1
            code = f"apu_auto_{auto_idx:04d}"
        key = code
        if key in result:
            auto_idx += 1
            key = f"{code}__{auto_idx}"
        result[key] = APUBreakdown(
            code=code,
            description=current_apu["description"],
            unit=unit,
            unit_price_total=round(float(total_value), 4),
            category=current_apu["category"],
            components=components,
            source=source,
        )
        current_apu = None

    header_seen = False
    
    max_col_needed = max(config.col_code, config.col_desc, config.col_qty, config.col_unit, config.col_price, config.col_subtotal, config.col_total, config.col_total_unit) + 1

    for row in _iter_rows(ws, max_col=max_col_needed):
        # Skip until we pass the header row "No. | PARTIDA | ...".
        if not header_seen:
            row_padded = row + (None,) * max_col_needed
            cell0 = _norm_text(row_padded[config.col_code]) if row else ""
            if cell0.lower().startswith("no"):
                header_seen = True
            continue

        if _is_blank_row(row):
            continue

        row_padded = row + (None,) * max_col_needed
        code_raw = row_padded[config.col_code]
        desc_raw = row_padded[config.col_desc]
        qty_raw = row_padded[config.col_qty]
        unit_raw = row_padded[config.col_unit]
        punit_raw = row_padded[config.col_price]
        valor_raw = row_padded[config.col_subtotal]
        total_raw = row_padded[config.col_total]
        total_unit_raw = row_padded[config.col_total_unit]
        description = _norm_text(desc_raw)
        qty = _to_float(qty_raw)
        unit_text = _norm_unit(unit_raw)
        punit = _to_float(punit_raw)
        valor = _to_float(valor_raw)
        total_value = _to_float(total_raw)
        total_unit = _norm_unit(total_unit_raw).lstrip("/").strip() if total_unit_raw else ""

        # Total row: closes the APU.
        if code_raw is None and not description and total_value is not None and total_unit:
            _close_current(total_value, total_unit)
            continue

        # Trailing absolute total (col 6 set, no total_unit) — ignored.
        if code_raw is None and not description and total_value is not None and not total_unit:
            continue

        if code_raw is not None:
            is_integer = isinstance(code_raw, (int, float)) and float(code_raw) == int(code_raw)
            code = _norm_code(code_raw)
            # Category banner: integer code, description, no qty/unit.
            if is_integer and description and qty is None and not unit_text:
                _close_current(None, None)
                current_category = description.upper()
                continue
            # New APU header (any code + description). Close any open APU first.
            if description:
                _close_current(None, None)
                current_apu = {
                    "code": code or "",
                    "description": description,
                    "unit": unit_text,
                    "category": current_category,
                    "components": [],
                }
                continue

        # Component row: code blank, description present.
        if current_apu is not None and description:
            comp_unit = unit_text
            comp_qty = qty if qty is not None else 0.0
            comp_price = punit if punit is not None else 0.0
            comp_subtotal = valor if valor is not None else round(comp_qty * comp_price, 4)
            ctype = _classify_component(description)
            current_apu["components"].append(
                APUComponent(
                    description=description,
                    quantity=round(float(comp_qty), 4),
                    unit=comp_unit,
                    unit_price=round(float(comp_price), 4),
                    subtotal=round(float(comp_subtotal), 4),
                    component_type=ctype,
                )
            )
            continue

    _close_current(None, None)
    return result


# ---------------------------------------------------------------------------
# Row iteration
# ---------------------------------------------------------------------------

def _iter_rows(ws: Worksheet, *, max_col: int) -> Iterator[tuple[Any, ...]]:
    for row in ws.iter_rows(values_only=True):
        yield tuple(row[:max_col])
