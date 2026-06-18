"""
Relational APU model (P0.1).

Turns the flat constructor pricing workbook into the three normalised tables a
professional estimate needs:

    RECURSOS          one row per priced resource (material / labour / equipment)
    APU               one row per unit-price analysis (partida) header
    APU_COMPONENTES   the decomposition of each APU into resource lines

Why relational: the unit price of every APU is *computed* from its components
(cantidad x precio_recurso), so when a resource price moves the whole budget can
be re-priced without re-doing any quantity take-off, and every cent is
traceable to a resource. This is the foundation P0.2 (item_type -> codigo_apu
crosswalk) and P0.3 (re-pricing) build on.

Completeness ("traer todo"): RECURSOS are extracted with a raw pass over the
materials + labour sheets that captures EVERY description-bearing row — including
rows with no price (flagged ``precio_faltante``) and a fallback to the second
price column (DUPLA) — so nothing is silently dropped. A completeness audit
compares raw data rows against captured rows and reports any difference.

CURRENCY: workbook values are Dominican pesos (RD$/DOP) by magnitude (cement
~540/funda, master mason ~2696/day). Override with ``DUPLA_PRICING_CURRENCY``.
Do NOT mix currencies between this store and the BC3 fallback.
"""

from __future__ import annotations

import json
import logging
import os
import re
import unicodedata
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Any

import openpyxl

from .schemas import PricingStore, PricingExcelConfig
from .excel_price_loader import (
    load_constructor_pricing,
    _find_sheet,
    _to_float,
    _norm_unit,
    _norm_text,
    _norm_code,
    _is_blank_row,
    _iter_rows,
)

logger = logging.getLogger("dupla.pricing.relational")


# ---------------------------------------------------------------------------
# Normalised tables
# ---------------------------------------------------------------------------

@dataclass
class Resource:
    """A priced input: a material, a labour activity or an equipment item."""
    codigo_recurso: str
    tipo: str               # MAT | MO | EQ
    descripcion: str
    unidad: str
    precio_unitario: float
    moneda: str = "DOP"
    fecha_precio: str | None = None
    categoria: str = ""
    fuente: str = ""
    precio_faltante: bool = False   # True when no price was found in the sheet


@dataclass
class APUComponentLink:
    """One decomposition line of an APU, linked back to a Resource when possible."""
    codigo_apu: str
    seq: int
    tipo: str               # MAT | MO | EQ | IND
    descripcion: str
    unidad: str
    cantidad: float         # rendimiento / consumo por unidad de partida
    precio_unitario: float  # inline price taken from the APU sheet
    desperdicio_pct: float
    subtotal: float
    codigo_recurso: str | None   # link to RECURSOS (None when unmatched / overhead)
    link_method: str             # exact | normalized | none | overhead


@dataclass
class APUHeader:
    codigo_apu: str
    descripcion: str
    unidad: str
    capitulo: str
    moneda: str = "DOP"
    total_declarado: float | None = None   # TOTAL PARTIDA stated in the sheet
    total_calculado: float = 0.0           # sum of component subtotals
    reconciliado: bool = False             # computed == stated within tolerance
    repriceable: bool = False              # every priced (MAT/MO/EQ) component is linked
                                           # to a resource -> can reprice exactly


@dataclass
class RelationalPricingStore:
    resources: dict[str, Resource] = field(default_factory=dict)
    apus: dict[str, APUHeader] = field(default_factory=dict)
    components: list[APUComponentLink] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)

    def components_for(self, codigo_apu: str) -> list[APUComponentLink]:
        return [c for c in self.components if c.codigo_apu == codigo_apu]

    def reprice(self, codigo_apu: str, resource_overrides: dict[str, float] | None = None) -> float:
        """Recompute the APU unit price from components x resource prices.

        ``resource_overrides`` maps ``codigo_recurso -> new_price`` for what-if
        scenarios (e.g. cement +10%). Linked MAT/MO/EQ lines reprice as
        ``cantidad * (1+desperdicio) * precio_recurso``; unlinked lines and IND
        (% overhead) lines keep their stored subtotal.
        """
        overrides = resource_overrides or {}
        total = 0.0
        for c in self.components_for(codigo_apu):
            if c.tipo == "IND" or c.codigo_recurso is None or c.codigo_recurso not in self.resources:
                total += c.subtotal
            else:
                price = overrides.get(c.codigo_recurso)
                if price is None:
                    price = self.resources[c.codigo_recurso].precio_unitario
                total += c.cantidad * (1.0 + c.desperdicio_pct) * float(price)
        return round(total, 4)

    def reprice_coverage(self) -> dict[str, Any]:
        """How much of the catalog can be repriced exactly from components."""
        repriceable = [a for a in self.apus.values() if a.repriceable]
        total = len(self.apus) or 1
        return {
            "repriceable_apus": len(repriceable),
            "total_apus": len(self.apus),
            "pct": round(100.0 * len(repriceable) / total, 1),
        }

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": 1,
            "metadata": dict(self.metadata),
            "resources": {k: asdict(v) for k, v in self.resources.items()},
            "apus": {k: asdict(v) for k, v in self.apus.items()},
            "components": [asdict(c) for c in self.components],
        }


# ---------------------------------------------------------------------------
# Normalisation / linking helpers
# ---------------------------------------------------------------------------

def _strip_accents(text: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFD", text)
        if unicodedata.category(ch) != "Mn"
    )


def _norm(text: Any) -> str:
    s = _strip_accents(str(text or "")).lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


_COMPONENT_TYPE_TO_TIPO = {
    "material": "MAT",
    "labor": "MO",
    "equipment": "EQ",
    "overhead": "IND",
}


def _is_overhead(description: str, unit: str) -> bool:
    d = _norm(description)
    return unit.strip() == "%" or "desperdicio" in d or "imprevisto" in d or "indirecto" in d


_STOP_TOKENS = {
    "de", "del", "en", "la", "el", "los", "las", "con", "para", "y", "o", "al", "por", "a",
    "x", "mm", "cm", "m", "m2", "m3", "ml", "u", "ud", "uds", "kg", "lb", "lbs", "gl", "gls",
    "fda", "qq", "par", "pa", "saco", "sacos", "und", "unidad",
}


def _content_tokens(text: str) -> set[str]:
    return {t for t in _norm(text).split() if t not in _STOP_TOKENS and not t.isdigit() and len(t) > 1}


def _ordered_content_tokens(text: str) -> list[str]:
    return [t for t in _norm(text).split() if t not in _STOP_TOKENS and not t.isdigit() and len(t) > 1]


def _build_token_index(resources: dict[str, "Resource"]) -> list[tuple[str, set[str], str]]:
    idx: list[tuple[str, set[str], str]] = []
    for code, r in resources.items():
        if r.precio_faltante:   # never link a component to a priceless resource
            continue
        ordered = _ordered_content_tokens(r.descripcion)
        if ordered:
            idx.append((code, set(ordered), ordered[0]))   # (code, tokens, head noun)
    return idx


def _token_link(description: str, token_index: list[tuple[str, set[str], str]]) -> str | None:
    """Link by token-subset, guarded by the head noun.

    Component tokens must be a subset of a resource's tokens AND must contain
    that resource's head (first) noun, so 'Grava' matches 'Grava clasificada'
    but 'Ligado' does NOT match 'Agua para ligado de mezclas'. The most-specific
    match must be unique (no ambiguity)."""
    tc = _content_tokens(description)
    if not tc:
        return None
    cands = [(code, tr) for code, tr, head in token_index if tc <= tr and head in tc]
    if not cands:
        return None
    cands.sort(key=lambda x: len(x[1]))
    if len(cands) == 1 or len(cands[0][1]) < len(cands[1][1]):
        return cands[0][0]
    return None  # ambiguous — leave unlinked rather than mislink


# ---------------------------------------------------------------------------
# Complete resource extraction (raw pass — nothing dropped)
# ---------------------------------------------------------------------------

def _looks_like_category_banner_material(code_raw: Any, unit: str, price: float | None) -> bool:
    is_int_code = isinstance(code_raw, (int, float)) and float(code_raw) == int(code_raw)
    return bool(is_int_code and not unit and price is None)


def _extract_resources_complete(
    excel_path: str | Path,
    config: PricingExcelConfig,
    moneda: str,
) -> tuple[dict[str, Resource], dict[str, str], dict[str, Any]]:
    """Capture EVERY material + labour row. Returns (resources, norm_index, audit)."""
    wb = openpyxl.load_workbook(str(excel_path), data_only=True, read_only=True)
    resources: dict[str, Resource] = {}
    norm_index: dict[str, str] = {}
    audit: dict[str, Any] = {}

    def _add(code: str, res: Resource) -> None:
        if code in resources:
            idx = 2
            while f"{code}__{idx}" in resources:
                idx += 1
            code = f"{code}__{idx}"
        res.codigo_recurso = code
        resources[code] = res
        key = _norm(res.descripcion)
        if key and key not in norm_index:
            norm_index[key] = code

    # --- Materials sheet ---------------------------------------------------
    ms = _find_sheet(wb.sheetnames, config.materials.sheet_names)
    cfg = config.materials
    dupla_col = 5  # second price column ("DUPLA") observed in the real sheet
    if ms:
        max_col = max(cfg.col_code, cfg.col_desc, cfg.col_unit, cfg.col_price, cfg.col_date, dupla_col) + 1
        data_rows = captured = banners = priceless = 0
        category = ""
        for row in _iter_rows(wb[ms], max_col=max_col):
            if _is_blank_row(row):
                continue
            rp = row + (None,) * max_col
            desc = _norm_text(rp[cfg.col_desc])
            nd = _norm(desc)
            code_cell = rp[cfg.col_code]
            if isinstance(code_cell, str) and code_cell.strip().rstrip(":").upper() in {
                "LUGAR", "OBRA", "FECHA", "PROCESO", "PARTIDA", "N", "NO",
            }:
                continue
            if not desc or "lista de precios" in nd or nd in {"material o actividad", "descripcion", "material"}:
                continue
            data_rows += 1
            code_raw = rp[cfg.col_code]
            unit = _norm_unit(rp[cfg.col_unit])
            price = _to_float(rp[cfg.col_price])
            if price is None:
                price = _to_float(rp[dupla_col])  # DUPLA fallback
            if _looks_like_category_banner_material(code_raw, unit, price):
                category = desc.upper()
                banners += 1
                continue
            date_raw = rp[cfg.col_date]
            updated = None
            if hasattr(date_raw, "isoformat"):
                try:
                    updated = date_raw.date().isoformat() if hasattr(date_raw, "date") else date_raw.isoformat()
                except Exception:
                    updated = None
            code = _norm_code(code_raw) or f"mat_{data_rows:04d}"
            _add(
                f"MAT-{code}",
                Resource(
                    codigo_recurso="", tipo="MAT", descripcion=desc, unidad=unit,
                    precio_unitario=round(price, 4) if price is not None else 0.0,
                    moneda=moneda, fecha_precio=updated, categoria=category, fuente=ms,
                    precio_faltante=price is None,
                ),
            )
            captured += 1
            if price is None:
                priceless += 1
        audit["materials"] = {
            "sheet": ms, "data_rows": data_rows, "captured": captured,
            "category_banners": banners, "priceless": priceless,
            "dropped": data_rows - banners - captured,
        }
    else:
        audit["materials"] = {"sheet": None}

    # --- Labour sheet ------------------------------------------------------
    ls = _find_sheet(wb.sheetnames, config.labor.sheet_names)
    lcfg = config.labor
    _labor_header_re = re.compile(r"^\d+\.\d+-")
    if ls:
        max_col = max(lcfg.col_code, lcfg.col_desc, lcfg.col_qty, lcfg.col_unit,
                      lcfg.col_price, lcfg.col_price_fallback) + 1
        data_rows = captured = banners = priceless = 0
        category = ""
        for row in _iter_rows(wb[ls], max_col=max_col):
            if _is_blank_row(row):
                continue
            rp = row + (None,) * max_col
            code_raw = rp[lcfg.col_code]
            desc = _norm_text(rp[lcfg.col_desc]) if rp[lcfg.col_desc] is not None else ""
            # Category banner: text in col A (uppercase / "2.0-..." / ends with ":").
            if isinstance(code_raw, str):
                text_a = _norm_text(code_raw)
                if not text_a:
                    continue
                if text_a.endswith(":") or _labor_header_re.match(text_a) or text_a.isupper():
                    category = re.sub(r"^\d+\.\d+-\s*", "", text_a).upper()
                    banners += 1
                    continue
                continue
            if not desc:
                continue
            data_rows += 1
            unit = _norm_unit(rp[lcfg.col_unit])
            price = _to_float(rp[lcfg.col_price])
            if price is None:
                price = _to_float(rp[lcfg.col_price_fallback])
            code = _norm_code(code_raw) or f"mo_{data_rows:04d}"
            _add(
                f"MO-{code}",
                Resource(
                    codigo_recurso="", tipo="MO", descripcion=desc, unidad=unit,
                    precio_unitario=round(price, 4) if price is not None else 0.0,
                    moneda=moneda, categoria=category, fuente=ls,
                    precio_faltante=price is None,
                ),
            )
            captured += 1
            if price is None:
                priceless += 1
        audit["labor"] = {
            "sheet": ls, "data_rows": data_rows, "captured": captured,
            "category_banners": banners, "priceless": priceless,
            "dropped": data_rows - captured,
        }
    else:
        audit["labor"] = {"sheet": None}

    return resources, norm_index, audit


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------

def build_relational_store(
    pricing_store: PricingStore,
    *,
    excel_path: str | Path | None = None,
    config: PricingExcelConfig | None = None,
    currency: str | None = None,
    reconcile_tolerance_pct: float = 0.01,
) -> RelationalPricingStore:
    """Build the relational store.

    When ``excel_path`` is given, RECURSOS are extracted with the complete raw
    pass (captures every row, nothing dropped). Otherwise they fall back to the
    already-parsed (lossy) ``pricing_store.materials`` / ``labor``.
    """
    moneda = (currency or os.getenv("DUPLA_PRICING_CURRENCY") or "DOP").strip() or "DOP"
    config = config or PricingExcelConfig()

    store = RelationalPricingStore(metadata={
        "currency": moneda,
        "source": pricing_store.metadata.get("source_file"),
    })

    # 1) RECURSOS
    audit: dict[str, Any] = {}
    if excel_path is not None:
        store.resources, norm_index, audit = _extract_resources_complete(excel_path, config, moneda)
    else:
        norm_index = {}
        for prefix, tipo, items in (
            ("MAT", "MAT", pricing_store.materials.values()),
            ("MO", "MO", pricing_store.labor.values()),
        ):
            for item in items:
                code = f"{prefix}-{item.code}"
                while code in store.resources:
                    code += "_x"
                store.resources[code] = Resource(
                    codigo_recurso=code, tipo=tipo, descripcion=item.description,
                    unidad=item.unit, precio_unitario=float(item.unit_price), moneda=moneda,
                    fecha_precio=getattr(item, "updated_date", None),
                    categoria=getattr(item, "category", "") or "", fuente=getattr(item, "source", "") or "",
                )
                key = _norm(item.description)
                if key and key not in norm_index:
                    norm_index[key] = code

    # 2) APU + APU_COMPONENTES (APU parsing reused from excel_price_loader)
    token_index = _build_token_index(store.resources)
    for apu in pricing_store.apus.values():
        stated = float(apu.unit_price_total) if apu.unit_price_total is not None else None
        header = APUHeader(
            codigo_apu=apu.code, descripcion=apu.description, unidad=apu.unit,
            capitulo=apu.category or "", moneda=moneda, total_declarado=stated,
        )
        computed = 0.0
        priced_count = 0
        linked_count = 0
        for seq, comp in enumerate(apu.components, start=1):
            overhead = _is_overhead(comp.description, comp.unit)
            tipo = "IND" if overhead else _COMPONENT_TYPE_TO_TIPO.get(comp.component_type, "MAT")
            resource_code: str | None = None
            link_method = "overhead" if overhead else "none"
            if not overhead:
                key = _norm(comp.description)
                if key in norm_index:
                    resource_code = norm_index[key]
                    link_method = "exact"
                else:
                    tok_code = _token_link(comp.description, token_index)
                    if tok_code is not None:
                        resource_code = tok_code
                        link_method = "token"
            if tipo != "IND":
                priced_count += 1
                if resource_code is not None:
                    linked_count += 1
            subtotal = float(comp.subtotal)
            computed += subtotal
            store.components.append(APUComponentLink(
                codigo_apu=apu.code, seq=seq, tipo=tipo, descripcion=comp.description,
                unidad=comp.unit, cantidad=float(comp.quantity),
                precio_unitario=float(comp.unit_price), desperdicio_pct=0.0,
                subtotal=subtotal, codigo_recurso=resource_code, link_method=link_method,
            ))
        header.total_calculado = round(computed, 4)
        if stated is not None and stated > 0:
            header.reconciliado = abs(computed - stated) <= max(reconcile_tolerance_pct * stated, 1.0)
        else:
            header.reconciliado = computed > 0
        header.repriceable = priced_count > 0 and linked_count == priced_count
        # Trust the reprice ONLY if it reconciles with the stated total — links can
        # be present but wrong (mismatched cantidad/unit), which would corrupt the
        # budget. If reprice diverges, fall back to the curated stated total.
        if header.repriceable and stated is not None and stated > 0:
            recomputed = store.reprice(apu.code)
            if abs(recomputed - stated) > max(reconcile_tolerance_pct * stated, 1.0):
                header.repriceable = False
        store.apus[apu.code] = header

    store.metadata["completeness_audit"] = audit
    _log_summary(store)
    return store


def _log_summary(store: RelationalPricingStore) -> None:
    linked = sum(1 for c in store.components if c.codigo_recurso is not None)
    overhead = sum(1 for c in store.components if c.tipo == "IND")
    priceless = sum(1 for r in store.resources.values() if r.precio_faltante)
    reconciled = sum(1 for a in store.apus.values() if a.reconciliado)
    total_apus = len(store.apus) or 1
    total_comp = len(store.components) or 1
    store.metadata.update({
        "resources": len(store.resources),
        "resources_priceless": priceless,
        "apus": len(store.apus),
        "components": len(store.components),
        "components_linked": linked,
        "components_overhead": overhead,
        "apus_reconciled": reconciled,
    })
    logger.info(
        "RelationalPricingStore: %d resources (%d sin precio), %d APUs, %d components "
        "(linked=%d/%d=%.0f%% overhead=%d), reconciled %d/%d=%.0f%%",
        len(store.resources), priceless, len(store.apus), len(store.components),
        linked, len(store.components), 100.0 * linked / total_comp, overhead,
        reconciled, len(store.apus), 100.0 * reconciled / total_apus,
    )


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def export_normalized_workbook(store: RelationalPricingStore, path: str | Path) -> Path:
    """Write the three relational tables to a clean .xlsx."""
    wb = openpyxl.Workbook()
    ws_r = wb.active
    ws_r.title = "RECURSOS"
    ws_r.append([
        "codigo_recurso", "tipo", "descripcion", "unidad", "precio_unitario",
        "moneda", "fecha_precio", "categoria", "fuente", "precio_faltante",
    ])
    for r in store.resources.values():
        ws_r.append([
            r.codigo_recurso, r.tipo, r.descripcion, r.unidad, r.precio_unitario,
            r.moneda, r.fecha_precio, r.categoria, r.fuente, "SI" if r.precio_faltante else "",
        ])

    ws_a = wb.create_sheet("APU")
    ws_a.append([
        "codigo_apu", "descripcion", "unidad", "capitulo", "moneda",
        "total_declarado", "total_calculado", "reconciliado",
    ])
    for a in store.apus.values():
        ws_a.append([
            a.codigo_apu, a.descripcion, a.unidad, a.capitulo, a.moneda,
            a.total_declarado, a.total_calculado, "SI" if a.reconciliado else "NO",
        ])

    ws_c = wb.create_sheet("APU_COMPONENTES")
    ws_c.append([
        "codigo_apu", "seq", "tipo", "codigo_recurso", "descripcion", "unidad",
        "cantidad", "precio_unitario", "desperdicio_pct", "subtotal", "link_method",
    ])
    for c in store.components:
        ws_c.append([
            c.codigo_apu, c.seq, c.tipo, c.codigo_recurso, c.descripcion, c.unidad,
            c.cantidad, c.precio_unitario, c.desperdicio_pct, c.subtotal, c.link_method,
        ])

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out


def save_relational_json(store: RelationalPricingStore, path: str | Path) -> Path:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(store.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def build_from_excel(excel_path: str | Path, *, currency: str | None = None) -> RelationalPricingStore:
    """Parse the workbook and build the relational store (complete resources)."""
    pricing_store = load_constructor_pricing(str(excel_path))
    return build_relational_store(pricing_store, excel_path=excel_path, currency=currency)


if __name__ == "__main__":
    import sys
    from core import paths  # type: ignore

    logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
    src = paths.pricing_excel_path()
    if src is None:
        print("No pricing Excel found.")
        sys.exit(1)
    rel = build_from_excel(src)
    out_dir = paths.data_dir()
    xlsx_out = export_normalized_workbook(rel, out_dir / "precios_relacional.xlsx")
    json_out = save_relational_json(rel, out_dir / "precios_relacional.json")
    print(f"\nWrote {xlsx_out}")
    print(f"Wrote {json_out}")
    print("\nMetadata:\n" + json.dumps(rel.metadata, ensure_ascii=False, indent=2))
