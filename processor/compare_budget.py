"""
Compare generated Dupla budget workbook against real PRES.xlsx baseline.

Usage:
    python compare_budget.py
    python compare_budget.py --generated "<path>" --real "./data/PRES.xlsx" --output-dir "<dir>"
"""

from __future__ import annotations

import argparse
import difflib
import json
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        cleaned = str(value).replace(",", ".").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0


def _normalize(text: str) -> str:
    lowered = text.lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
        "�": "a",
    }
    for src, dst in replacements.items():
        lowered = lowered.replace(src, dst)
    return lowered


def _normalize_code(code: Any) -> str:
    """Normalize code text for resilient matching across workbook formats."""
    raw = _safe_str(code)
    if not raw:
        return ""
    # Keep only alphanumeric chars and normalize casing.
    return re.sub(r"[^A-Za-z0-9]", "", raw).upper()


def _index_by_normalized_code(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    indexed: dict[str, dict[str, Any]] = {}
    for row in rows:
        normalized = _normalize_code(row.get("code"))
        if not normalized:
            continue
        # Prefer the row with the largest amount when normalized codes collide.
        current = indexed.get(normalized)
        if current is None or _safe_float(row.get("amount")) > _safe_float(current.get("amount")):
            indexed[normalized] = row
    return indexed


def _load_budget_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        code = _safe_str(row[0] if len(row) > 0 else None)
        nat = _safe_str(row[1] if len(row) > 1 else None)
        unit = _safe_str(row[2] if len(row) > 2 else None)
        summary = _safe_str(row[3] if len(row) > 3 else None)
        qty = _safe_float(row[4] if len(row) > 4 else None)
        price = _safe_float(row[5] if len(row) > 5 else None)
        amount = _safe_float(row[6] if len(row) > 6 else None)
        if not any((code, nat, unit, summary)):
            continue
        rows.append(
            {
                "code": code,
                "nat": nat,
                "unit": unit,
                "summary": summary,
                "quantity": qty,
                "price": price,
                "amount": amount,
            }
        )
    return rows


def _is_partida(row: dict[str, Any]) -> bool:
    return "partida" in _normalize(str(row.get("nat", "")))


def _is_chapter(row: dict[str, Any]) -> bool:
    return "cap" in _normalize(str(row.get("nat", "")))


def _discipline_tags(text: str) -> set[str]:
    normalized = _normalize(text)
    tags: dict[str, tuple[str, ...]] = {
        "preliminares": ("prelimin",),
        "movimiento_tierra": ("movimiento de tierra", "excav", "relleno"),
        "hormigon_armado": ("hormigon", "hormigon armado", "concreto"),
        "acero_refuerzo": ("acero", "refuerzo", "varilla"),
        "muros_divisiones": ("muro", "bloque", "division"),
        "panete_revestimiento": ("panete", "pañete", "fraguache", "revest"),
        "pisos": ("piso", "porcelanato", "ceram", "zocalo"),
        "escaleras": ("escalera", "escalon"),
        "puertas": ("puerta",),
        "ventanas": ("ventana", "vidrio"),
        "ebanisteria": ("ebanister", "closet", "gabinete"),
        "electrico": ("electrico", "eléctrico", "luminaria", "tomacorr", "panel"),
        "sanitario": ("sanitario", "inodoro", "lavamanos", "plomer", "drenaje"),
        "pintura": ("pintura", "sellador", "imperme"),
        "techos_cubierta": ("techo", "cubierta"),
        "miscelaneos": ("miscelaneo", "miscelaneo"),
        "equipos_electricos": ("equipos electric",),
        "herreria": ("verja", "baranda", "herreria"),
        "impermeabilizacion": ("impermeabil",),
        "acabados": ("terminacion", "acabado"),
        "gastos_generales": ("gastos", "indirectos", "supervision"),
    }
    found: set[str] = set()
    for tag, hints in tags.items():
        if any(hint in normalized for hint in hints):
            found.add(tag)
    return found


# Tags that belong to each pipeline discipline for honest coverage metrics.
DISCIPLINE_FILTER_TAGS: dict[str, set[str]] = {
    "arquitectura": {
        "muros_divisiones", "panete_revestimiento", "pisos", "escaleras",
        "puertas", "ventanas", "ebanisteria", "pintura", "techos_cubierta",
        "herreria", "impermeabilizacion", "acabados",
    },
    "estructura": {
        "movimiento_tierra", "hormigon_armado", "acero_refuerzo",
    },
    "electrico": {
        "electrico", "equipos_electricos",
    },
    "sanitario": {
        "sanitario",
    },
}


def _filter_partidas_by_discipline(
    partidas: list[dict[str, Any]],
    discipline: str,
    all_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Return partidas whose chapter context matches the given discipline.

    Uses the chapter heading above each partida to determine discipline.
    Falls back to partida-level tags when chapter context is unavailable.
    """
    allowed_tags = DISCIPLINE_FILTER_TAGS.get(discipline)
    if not allowed_tags:
        return partidas

    # Build a chapter-context lookup: for each partida, find its parent chapter's tags.
    chapter_context: dict[int, set[str]] = {}
    current_chapter_tags: set[str] = set()
    for idx, row in enumerate(all_rows):
        if _is_chapter(row):
            current_chapter_tags = _discipline_tags(_safe_str(row.get("summary")))
        chapter_context[idx] = current_chapter_tags

    # Map partidas to their row index for chapter context lookup.
    partida_indices: dict[int, dict[str, Any]] = {}
    partida_idx = 0
    for idx, row in enumerate(all_rows):
        if _is_partida(row) and partida_idx < len(partidas):
            partida_indices[idx] = partidas[partida_idx]
            partida_idx += 1

    filtered: list[dict[str, Any]] = []
    for idx, partida in partida_indices.items():
        ctx_tags = chapter_context.get(idx, set())
        partida_tags = _row_family_tags(partida)
        combined = ctx_tags | partida_tags
        if combined & allowed_tags:
            filtered.append(partida)

    return filtered


def _line_precision(real_value: float, gen_value: float) -> float:
    if real_value == 0 and gen_value == 0:
        return 1.0
    if real_value == 0:
        return 0.0
    score = 1.0 - abs(gen_value - real_value) / abs(real_value)
    return max(0.0, min(1.0, score))


def _mean(values: list[float]) -> float:
    if not values:
        return 0.0
    return sum(values) / len(values)


def _md_cell(value: Any) -> str:
    text = _safe_str(value).replace("|", "\\|").replace("\n", " ")
    return text


def _tokenize_summary(text: str) -> set[str]:
    normalized = _normalize(text)
    tokens = set(re.findall(r"[a-z0-9]+", normalized))
    stopwords = {
        "de", "la", "el", "y", "en", "con", "para", "por", "del", "los", "las",
        "un", "una", "al", "se", "a", "o", "que", "nivel", "obra",
    }
    return {token for token in tokens if len(token) >= 3 and token not in stopwords}


def _row_family_tags(row: dict[str, Any]) -> set[str]:
    summary = _safe_str(row.get("summary"))
    nat = _safe_str(row.get("nat"))
    return _discipline_tags(f"{summary} {nat}".strip())


def _candidate_mapping_score(generated: dict[str, Any], real: dict[str, Any]) -> float:
    generated_summary = _normalize(_safe_str(generated.get("summary")))
    real_summary = _normalize(_safe_str(real.get("summary")))
    if not generated_summary or not real_summary:
        return 0.0

    seq_similarity = difflib.SequenceMatcher(None, generated_summary, real_summary).ratio()

    generated_tokens = _tokenize_summary(generated_summary)
    real_tokens = _tokenize_summary(real_summary)
    if generated_tokens or real_tokens:
        token_overlap = len(generated_tokens & real_tokens) / max(1, len(generated_tokens | real_tokens))
    else:
        token_overlap = 0.0

    generated_unit = _safe_str(generated.get("unit")).lower()
    real_unit = _safe_str(real.get("unit")).lower()
    unit_bonus = 0.12 if generated_unit and real_unit and generated_unit == real_unit else 0.0

    generated_families = _row_family_tags(generated)
    real_families = _row_family_tags(real)
    family_bonus = 0.12 if generated_families & real_families else 0.0

    raw_score = 0.58 * seq_similarity + 0.30 * token_overlap + unit_bonus + family_bonus
    return min(1.0, raw_score)


def _semantic_similarity_metrics(
    generated_partidas: list[dict[str, Any]],
    real_partidas: list[dict[str, Any]],
) -> dict[str, float]:
    """Compute lightweight semantic overlap proxies based on summary text similarity."""
    real_by_unit: dict[str, list[str]] = {}
    for row in real_partidas:
        unit = _safe_str(row.get("unit")).lower()
        summary = _normalize(_safe_str(row.get("summary")))
        if not summary:
            continue
        real_by_unit.setdefault(unit, []).append(summary)

    best_scores: list[float] = []
    for row in generated_partidas:
        unit = _safe_str(row.get("unit")).lower()
        generated_summary = _normalize(_safe_str(row.get("summary")))
        if not generated_summary:
            continue

        candidates = real_by_unit.get(unit) or [
            summary for summaries in real_by_unit.values() for summary in summaries
        ]
        if not candidates:
            continue

        best = max(
            difflib.SequenceMatcher(None, generated_summary, candidate).ratio()
            for candidate in candidates
        )
        best_scores.append(best)

    if not best_scores:
        return {
            "semantic_avg_best_similarity": 0.0,
            "semantic_match_rate_60": 0.0,
            "semantic_match_rate_70": 0.0,
        }

    rate60 = sum(1 for score in best_scores if score >= 0.60) / len(best_scores)
    rate70 = sum(1 for score in best_scores if score >= 0.70) / len(best_scores)
    return {
        "semantic_avg_best_similarity": 100.0 * _mean(best_scores),
        "semantic_match_rate_60": 100.0 * rate60,
        "semantic_match_rate_70": 100.0 * rate70,
    }


def _map_generated_to_real_codes(
    generated_partidas: list[dict[str, Any]],
    real_partidas: list[dict[str, Any]],
    *,
    min_similarity: float = 0.52,
) -> dict[str, Any]:
    """Infer a PRES-code mapping from generated lines using summary/unit similarity.

    Returns a one-to-one mapping (best generated candidate per real code).
    """
    real_by_unit: dict[str, list[dict[str, Any]]] = {}
    for row in real_partidas:
        unit = _safe_str(row.get("unit")).lower()
        real_by_unit.setdefault(unit, []).append(row)

    candidate_by_real_code: dict[str, dict[str, Any]] = {}

    for generated in generated_partidas:
        generated_summary = _safe_str(generated.get("summary"))
        if not generated_summary:
            continue

        unit = _safe_str(generated.get("unit")).lower()
        candidates = real_by_unit.get(unit) or real_partidas
        generated_families = _row_family_tags(generated)
        if generated_families:
            family_candidates = [
                candidate for candidate in candidates if _row_family_tags(candidate) & generated_families
            ]
            if family_candidates:
                candidates = family_candidates
        if not candidates:
            continue

        best_row: dict[str, Any] | None = None
        best_score = 0.0
        for real in candidates:
            score = _candidate_mapping_score(generated, real)
            if score > best_score:
                best_score = score
                best_row = real

        if best_row is None or best_score < min_similarity:
            continue

        real_code = _safe_str(best_row.get("code"))
        if not real_code:
            continue

        previous = candidate_by_real_code.get(real_code)
        current = {
            "real": best_row,
            "generated": generated,
            "score": best_score,
        }
        if previous is None or current["score"] > previous["score"]:
            candidate_by_real_code[real_code] = current

    mapped_pairs = list(candidate_by_real_code.values())
    qty_precisions = [
        _line_precision(pair["real"]["quantity"], pair["generated"]["quantity"])
        for pair in mapped_pairs
    ]
    price_precisions = [
        _line_precision(pair["real"]["price"], pair["generated"]["price"])
        for pair in mapped_pairs
    ]

    mapped_coverage = 100.0 * (len(mapped_pairs) / len(real_partidas)) if real_partidas else 0.0

    top_pairs: list[dict[str, Any]] = []
    for pair in sorted(mapped_pairs, key=lambda item: item["score"], reverse=True)[:20]:
        real = pair["real"]
        generated = pair["generated"]
        top_pairs.append(
            {
                "real_code": _safe_str(real.get("code")),
                "generated_code": _safe_str(generated.get("code")),
                "score": round(100.0 * float(pair["score"]), 2),
                "real_summary": _safe_str(real.get("summary"))[:120],
                "generated_summary": _safe_str(generated.get("summary"))[:120],
                "real_unit": _safe_str(real.get("unit")),
                "generated_unit": _safe_str(generated.get("unit")),
            }
        )

    return {
        "mapped_pairs": mapped_pairs,
        "mapped_coverage_pres_code": mapped_coverage,
        "mapped_qty_accuracy": 100.0 * _mean(qty_precisions),
        "mapped_price_accuracy": 100.0 * _mean(price_precisions),
        "mapped_codes": sorted(candidate_by_real_code),
        "mapped_count": len(mapped_pairs),
        "mapped_top_pairs": top_pairs,
        "mapped_min_similarity": min_similarity,
    }


def analyze_budget_pair(
    generated_path: Path,
    real_path: Path,
    *,
    discipline_filter: str | None = None,
) -> dict[str, Any]:
    """
    Métricas compartidas entre el informe .txt y el informe Markdown.
    """
    generated_rows = _load_budget_rows(generated_path)
    real_rows = _load_budget_rows(real_path)

    generated_partidas = [row for row in generated_rows if _is_partida(row)]
    real_partidas = [row for row in real_rows if _is_partida(row)]
    generated_chapters = [row for row in generated_rows if _is_chapter(row)]
    real_chapters = [row for row in real_rows if _is_chapter(row)]

    generated_by_code: dict[str, dict[str, Any]] = {
        _safe_str(row["code"]): row for row in generated_partidas if _safe_str(row["code"])
    }
    real_by_code: dict[str, dict[str, Any]] = {
        _safe_str(row["code"]): row for row in real_partidas if _safe_str(row["code"])
    }
    generated_by_code_normalized = _index_by_normalized_code(generated_partidas)
    real_by_code_normalized = _index_by_normalized_code(real_partidas)

    matching_codes_exact = sorted(set(real_by_code) & set(generated_by_code))
    matching_codes_normalized = sorted(set(real_by_code_normalized) & set(generated_by_code_normalized))
    matching_codes = matching_codes_normalized

    real_only_codes = sorted(set(real_by_code) - set(generated_by_code))
    generated_only_codes = sorted(set(generated_by_code) - set(real_by_code))

    real_only_codes_normalized = sorted(
        set(real_by_code_normalized) - set(generated_by_code_normalized)
    )
    generated_only_codes_normalized = sorted(
        set(generated_by_code_normalized) - set(real_by_code_normalized)
    )

    qty_precisions = [
        _line_precision(
            real_by_code_normalized[code]["quantity"],
            generated_by_code_normalized[code]["quantity"],
        )
        for code in matching_codes
    ]
    price_precisions = [
        _line_precision(
            real_by_code_normalized[code]["price"],
            generated_by_code_normalized[code]["price"],
        )
        for code in matching_codes
    ]

    generated_disciplines: set[str] = set()
    for row in generated_rows:
        generated_disciplines.update(_discipline_tags(row["summary"]))
    real_disciplines: set[str] = set()
    for row in real_rows:
        real_disciplines.update(_discipline_tags(row["summary"]))

    top20_real = sorted(real_partidas, key=lambda row: row["amount"], reverse=True)[:20]

    generated_non_empty_price = sum(1 for row in generated_partidas if row["price"] > 0)
    generated_non_empty_amount = sum(1 for row in generated_partidas if row["amount"] > 0)
    real_total = sum(row["amount"] for row in real_partidas)
    generated_total = sum(row["amount"] for row in generated_partidas)

    coverage_exact = 100.0 * (len(matching_codes_exact) / len(real_by_code)) if real_by_code else 0.0
    coverage = (
        100.0 * (len(matching_codes_normalized) / len(real_by_code_normalized))
        if real_by_code_normalized
        else 0.0
    )
    qty_accuracy = 100.0 * _mean(qty_precisions)
    price_accuracy = 100.0 * _mean(price_precisions)

    amount_deltas: list[dict[str, Any]] = []
    for code in matching_codes:
        r = real_by_code_normalized[code]
        g = generated_by_code_normalized[code]
        amount_deltas.append(
            {
                "code": _safe_str(r.get("code")) or code,
                "normalized_code": code,
                "real_amount": r["amount"],
                "gen_amount": g["amount"],
                "delta": g["amount"] - r["amount"],
                "real_qty": r["quantity"],
                "gen_qty": g["quantity"],
                "summary": _safe_str(r["summary"])[:120],
            }
        )
    amount_deltas.sort(key=lambda item: abs(float(item["delta"])), reverse=True)
    semantic = _semantic_similarity_metrics(generated_partidas, real_partidas)
    mapped = _map_generated_to_real_codes(generated_partidas, real_partidas, min_similarity=0.52)

    # Discipline-filtered metrics: compare only against the relevant subset of PRES.
    filtered_metrics: dict[str, Any] = {
        "discipline_filter": discipline_filter,
        "filtered_real_count": len(real_partidas),
        "filtered_semantic": {},
        "filtered_mapped": {},
    }
    if discipline_filter:
        filtered_real = _filter_partidas_by_discipline(real_partidas, discipline_filter, real_rows)
        filtered_metrics["filtered_real_count"] = len(filtered_real)
        if filtered_real:
            filtered_metrics["filtered_semantic"] = _semantic_similarity_metrics(
                generated_partidas, filtered_real,
            )
            filtered_metrics["filtered_mapped"] = _map_generated_to_real_codes(
                generated_partidas, filtered_real, min_similarity=0.35,
            )
        filtered_metrics["filtered_coverage_theoretical"] = (
            100.0 * len(generated_partidas) / len(filtered_real) if filtered_real else 0.0
        )

    return {
        "generated_rows": generated_rows,
        "real_rows": real_rows,
        "generated_partidas": generated_partidas,
        "real_partidas": real_partidas,
        "generated_chapters": generated_chapters,
        "real_chapters": real_chapters,
        "generated_by_code": generated_by_code,
        "real_by_code": real_by_code,
        "generated_by_code_normalized": generated_by_code_normalized,
        "real_by_code_normalized": real_by_code_normalized,
        "matching_codes": matching_codes,
        "matching_codes_exact": matching_codes_exact,
        "matching_codes_normalized": matching_codes_normalized,
        "real_only_codes": real_only_codes,
        "generated_only_codes": generated_only_codes,
        "real_only_codes_normalized": real_only_codes_normalized,
        "generated_only_codes_normalized": generated_only_codes_normalized,
        "qty_precisions": qty_precisions,
        "price_precisions": price_precisions,
        "generated_disciplines": generated_disciplines,
        "real_disciplines": real_disciplines,
        "top20_real": top20_real,
        "generated_non_empty_price": generated_non_empty_price,
        "generated_non_empty_amount": generated_non_empty_amount,
        "real_total": real_total,
        "generated_total": generated_total,
        "coverage": coverage,
        "coverage_exact": coverage_exact,
        "qty_accuracy": qty_accuracy,
        "price_accuracy": price_accuracy,
        "amount_deltas": amount_deltas,
        **semantic,
        **mapped,
        **filtered_metrics,
    }


def build_comparison_markdown(
    generated_path: Path,
    real_path: Path,
    *,
    title: str,
    run_date: str,
    run_tag: str,
    notes: str = "",
    max_list_codes: int = 80,
    max_delta_rows: int = 25,
) -> str:
    """
    Informe en Markdown para carpetas de comparación por proyecto/corrida.
    """
    stats = analyze_budget_pair(generated_path, real_path)
    lines: list[str] = [
        f"# {title}",
        "",
        f"- **Fecha de corrida:** {run_date}",
        f"- **Etiqueta de corrida:** `{run_tag}`",
        f"- **Generado (Dupla):** `{generated_path}`",
        f"- **Referencia (PRES):** `{real_path}`",
        "",
        "## Contexto y limitaciones",
        "",
        notes.strip() or (
            "- Esta comparación asume el mismo layout Presto en la primera hoja (filas desde la 4). "
            "- Requiere validación manual si el PRES usa otra hoja o formato."
        ),
        "",
        "## Resumen ejecutivo",
        "",
        "| Métrica | Generado | PRES (real) |",
        "| --- | ---: | ---: |",
        f"| Partidas | {len(stats['generated_partidas'])} | {len(stats['real_partidas'])} |",
        f"| Capítulos (filas Nat) | {len(stats['generated_chapters'])} | {len(stats['real_chapters'])} |",
        f"| Códigos coincidentes | {len(stats['matching_codes'])} | — |",
        f"| Códigos coincidentes (exactos) | {len(stats['matching_codes_exact'])} | — |",
        f"| Códigos coincidentes (normalizados) | {len(stats['matching_codes_normalized'])} | — |",
        f"| Cobertura exacta (código crudo) | {stats['coverage_exact']:.2f}% | — |",
        f"| Cobertura códigos PRES con equivalente generado | {stats['coverage']:.2f}% | — |",
        f"| Precisión cantidad (solo códigos coincidentes) | {stats['qty_accuracy']:.2f}% | — |",
        f"| Precisión precio unitario (solo coincidentes) | {stats['price_accuracy']:.2f}% | — |",
        f"| Similitud semántica promedio (resumen) | {stats['semantic_avg_best_similarity']:.2f}% | — |",
        f"| Match semántico >= 60% | {stats['semantic_match_rate_60']:.2f}% | — |",
        f"| Match semántico >= 70% | {stats['semantic_match_rate_70']:.2f}% | — |",
        f"| Cobertura mapeada a código PRES (sim>=60%) | {stats['mapped_coverage_pres_code']:.2f}% | — |",
        f"| Precisión cantidad mapeada | {stats['mapped_qty_accuracy']:.2f}% | — |",
        f"| Precisión precio mapeada | {stats['mapped_price_accuracy']:.2f}% | — |",
        f"| Suma Importe (ImpPres) | {stats['generated_total']:,.2f} | {stats['real_total']:,.2f} |",
        f"| Delta (generado − real) | {stats['generated_total'] - stats['real_total']:,.2f} | — |",
        "",
        "### Completitud de precios en generado",
        "",
        f"- Filas partida con PrPres > 0: **{stats['generated_non_empty_price']}** / {len(stats['generated_partidas'])}",
        f"- Filas partida con ImpPres > 0: **{stats['generated_non_empty_amount']}** / {len(stats['generated_partidas'])}",
        "",
        "## Disciplinas (heurística por texto del resumen)",
        "",
        f"- Etiquetas presentes en PRES y no detectadas en generado: **{', '.join(sorted(stats['real_disciplines'] - stats['generated_disciplines'])) or '—'}**",
        "",
        "## Mayores diferencias de importe (códigos en ambos)",
        "",
        "| Código | ImpPres real | ImpPres gen | Delta | Cant. real | Cant. gen | Resumen (PRES) |",
        "| --- | ---: | ---: | ---: | ---: | ---: | --- |",
    ]
    for row in stats["amount_deltas"][:max_delta_rows]:
        lines.append(
            "| {code} | {ra:,.2f} | {ga:,.2f} | {d:,.2f} | {rq} | {gq} | {s} |".format(
                code=_md_cell(row["code"]),
                ra=row["real_amount"],
                ga=row["gen_amount"],
                d=row["delta"],
                rq=row["real_qty"],
                gq=row["gen_qty"],
                s=_md_cell(row["summary"]),
            )
        )
    lines.extend(
        [
            "",
            "## Códigos solo en PRES (no aparecen en generado)",
            "",
        ]
    )
    roc = stats["real_only_codes"]
    if not roc:
        lines.append("_Ninguno._")
    else:
        shown = roc[:max_list_codes]
        lines.extend(f"- `{c}`" for c in shown)
        if len(roc) > max_list_codes:
            lines.append(f"- … y **{len(roc) - max_list_codes}** más.")
    lines.extend(["", "## Códigos solo en generado (no están en PRES)", ""])
    goc = stats["generated_only_codes"]
    if not goc:
        lines.append("_Ninguno._")
    else:
        shown = goc[:max_list_codes]
        lines.extend(f"- `{c}`" for c in shown)
        if len(goc) > max_list_codes:
            lines.append(f"- … y **{len(goc) - max_list_codes}** más.")

    lines.extend(["", "## Top 20 partidas PRES por importe vs generado", ""])
    for row in stats["top20_real"]:
        code = _safe_str(row["code"])
        gen = stats["generated_by_code"].get(code)
        if gen is None:
            lines.append(
                f"- **{code}**: real **{row['amount']:,.2f}** — generado _no encontrado_ — {_md_cell(row['summary'])}"
            )
            continue
        qty_score = 100.0 * _line_precision(row["quantity"], gen["quantity"])
        price_score = 100.0 * _line_precision(row["price"], gen["price"])
        lines.append(
            f"- **{code}**: real **{row['amount']:,.2f}** | gen **{gen['amount']:,.2f}** | "
            f"precisión cant. {qty_score:.1f}% | precio {price_score:.1f}%"
        )

    lines.extend(["", "## Top mapeos inferidos (similitud resumen)", ""])
    if not stats["mapped_top_pairs"]:
        lines.append("_Sin mapeos inferidos con umbral actual._")
    else:
        lines.append("| Código PRES | Código generado | Similitud | U. PRES | U. gen | Resumen PRES | Resumen generado |")
        lines.append("| --- | --- | ---: | --- | --- | --- | --- |")
        for row in stats["mapped_top_pairs"]:
            lines.append(
                "| {rc} | {gc} | {sc:.2f}% | {ru} | {gu} | {rs} | {gs} |".format(
                    rc=_md_cell(row["real_code"]),
                    gc=_md_cell(row["generated_code"]),
                    sc=float(row["score"]),
                    ru=_md_cell(row["real_unit"]),
                    gu=_md_cell(row["generated_unit"]),
                    rs=_md_cell(row["real_summary"]),
                    gs=_md_cell(row["generated_summary"]),
                )
            )

    return "\n".join(lines) + "\n"


def build_comparison_report(generated_path: Path, real_path: Path, output_dir: Path) -> Path:
    stats = analyze_budget_pair(generated_path, real_path)
    matching_codes = stats["matching_codes"]
    qty_precisions = stats["qty_precisions"]
    price_precisions = stats["price_precisions"]
    generated_disciplines = stats["generated_disciplines"]
    real_disciplines = stats["real_disciplines"]
    top20_real = stats["top20_real"]
    generated_partidas = stats["generated_partidas"]
    real_partidas = stats["real_partidas"]
    generated_chapters = stats["generated_chapters"]
    real_chapters = stats["real_chapters"]
    generated_by_code = stats["generated_by_code"]
    real_by_code = stats["real_by_code"]
    generated_non_empty_price = stats["generated_non_empty_price"]
    generated_non_empty_amount = stats["generated_non_empty_amount"]
    real_total = stats["real_total"]
    generated_total = stats["generated_total"]
    coverage = stats["coverage"]
    qty_accuracy = stats["qty_accuracy"]
    price_accuracy = stats["price_accuracy"]

    lines = [
        "COMPARISON REPORT - DUPLA VS PRES.xlsx",
        "",
        f"Generated workbook: {generated_path}",
        f"Real workbook: {real_path}",
        "",
        "1) High-level counts",
        f"- Partidas generated: {len(generated_partidas)} | real: {len(real_partidas)} (expected ~1565)",
        f"- Chapters generated: {len(generated_chapters)} | real: {len(real_chapters)} (expected ~296)",
        (
            f"- Disciplines covered: {len(generated_disciplines)} | "
            f"real: {len(real_disciplines)} (expected ~21)"
        ),
        "",
        "2) Price/amount completeness in generated",
        f"- Rows with PrPres > 0: {generated_non_empty_price}/{len(generated_partidas)}",
        f"- Rows with ImpPres > 0: {generated_non_empty_amount}/{len(generated_partidas)}",
        "",
        "3) Matching quality by code",
        f"- Matching codes (normalized): {len(matching_codes)}",
        f"- Matching codes (exact raw): {len(stats['matching_codes_exact'])}",
        f"- Coverage (exact raw): {stats['coverage_exact']:.2f}%",
        f"- Coverage of real partidas by generated equivalent: {coverage:.2f}%",
        f"- Quantity precision (only matched codes): {qty_accuracy:.2f}%",
        f"- Price precision (only matched codes): {price_accuracy:.2f}%",
        f"- Semantic avg best similarity (summary): {stats['semantic_avg_best_similarity']:.2f}%",
        f"- Semantic match rate >= 60%: {stats['semantic_match_rate_60']:.2f}%",
        f"- Semantic match rate >= 70%: {stats['semantic_match_rate_70']:.2f}%",
        f"- Mapped PRES code coverage (sim >= {100.0*stats['mapped_min_similarity']:.0f}%): {stats['mapped_coverage_pres_code']:.2f}%",
        f"- Mapped quantity precision: {stats['mapped_qty_accuracy']:.2f}%",
        f"- Mapped price precision: {stats['mapped_price_accuracy']:.2f}%",
        "",
        "4) Totals",
        f"- Generated total (sum ImpPres): {generated_total:,.2f}",
        f"- Real total (sum ImpPres): {real_total:,.2f} (reference target: 4,404,786 USD)",
        f"- Delta generated-real: {generated_total - real_total:,.2f}",
        "",
        "5) Missing disciplines (in generated vs real)",
    ]
    missing_disciplines = sorted(real_disciplines - generated_disciplines)
    if missing_disciplines:
        lines.extend(f"- {discipline}" for discipline in missing_disciplines)
    else:
        lines.append("- None")

    lines.extend(
        [
            "",
            "6) Top 20 real partidas by amount vs generated",
        ]
    )
    for row in top20_real:
        code = row["code"]
        gen = generated_by_code.get(code)
        if gen is None:
            lines.append(
                f"- {code}: real_amount={row['amount']:,.2f} | generated=NOT_FOUND | summary={row['summary']}"
            )
            continue
        qty_score = 100.0 * _line_precision(row["quantity"], gen["quantity"])
        price_score = 100.0 * _line_precision(row["price"], gen["price"])
        lines.append(
            (
                f"- {code}: real_amount={row['amount']:,.2f} | gen_amount={gen['amount']:,.2f} | "
                f"qty_precision={qty_score:.2f}% | price_precision={price_score:.2f}%"
            )
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    report_path = output_dir / "comparison_report.txt"
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


def _resolve_defaults() -> tuple[Path, Path, Path]:
    repo_root = Path(__file__).resolve().parent
    run_summary = Path(r"C:\Users\chris\Downloads\archivos dupla\dwg\run_summary.json")
    if run_summary.exists():
        payload = json.loads(run_summary.read_text(encoding="utf-8"))
        generated = Path(payload["budget_excel"])
        output_dir = generated.parent
        real = repo_root / "data" / "PRES.xlsx"
        return generated, real, output_dir
    return (
        repo_root / "output" / "dupla_budget_ready_full.xlsx",
        repo_root / "data" / "PRES.xlsx",
        repo_root / "output",
    )


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Compare generated budget vs PRES.xlsx")
    defaults_generated, defaults_real, defaults_output = _resolve_defaults()
    parser.add_argument("--generated", default=str(defaults_generated))
    parser.add_argument("--real", default=str(defaults_real))
    parser.add_argument("--output-dir", default=str(defaults_output))
    return parser


def main() -> None:
    args = _build_arg_parser().parse_args()
    generated = Path(args.generated).resolve()
    real = Path(args.real).resolve()
    output_dir = Path(args.output_dir).resolve()

    if not generated.exists():
        raise FileNotFoundError(f"Generated workbook not found: {generated}")
    if not real.exists():
        raise FileNotFoundError(f"Real workbook not found: {real}")

    report_path = build_comparison_report(generated, real, output_dir)
    print(f"Comparison report written to: {report_path}")


if __name__ == "__main__":
    main()
