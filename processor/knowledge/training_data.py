"""
Training dataset extraction helpers from real completed budgets.

Supports loading from:
- A single PRES.xlsx file (backward-compatible API).
- A folder of XLSX files (multi-project corpus from Dropbox).
"""

from __future__ import annotations

from dataclasses import dataclass
import json
import logging
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger("dupla.training_data")


@dataclass
class TrainingPair:
    input_item_type: str
    input_unit: str
    input_context: str
    output_bc3_code: str
    output_description: str
    output_unit: str
    output_quantity: float
    output_price: float
    source: str = "PRES.xlsx"


@dataclass
class LevelTemplate:
    template_id: str
    levels: list[str]
    discipline: str
    item_codes: list[str]
    item_types: list[str]
    item_count: int
    signature: str


_ITEM_TYPE_HINTS: list[tuple[tuple[str, ...], str]] = [
    (("excavac", "rellen", "terracer", "compactac", "movimiento de tierra", "bote"), "generic_construction_item"),
    (("zapata", "platea", "fundacion"), "footing"),
    (("columna",), "column"),
    (("viga",), "beam"),
    (("losa",), "slab"),
    (("panete", "pañete", "fraguache"), "wall_finish_plaster"),
    (("muro bloques", "muro de bloque", "muro h.a.", "muro"), "wall"),
    (("porcelanato", "ceramica", "cerámica", "piso"), "floor_finish"),
    (("puerta",), "door_count"),
    (("ventana",), "window_count"),
    (("pintura",), "wall_finish_paint"),
    (("sanitario", "inodoro", "lavamanos", "ducha"), "wet_area_fixture_count"),
    (("electrico", "eléctrico", "tomacorriente", "interruptor", "luminaria"), "fixture_count"),
]

_LEVEL_PATTERNS = [
    re.compile(r"\bsemisotano\b", re.IGNORECASE),
    re.compile(r"\bnivel\s+\d+\b", re.IGNORECASE),
    re.compile(r"\btecho\b", re.IGNORECASE),
    re.compile(r"\bmiscelaneos\b", re.IGNORECASE),
    re.compile(r"\bequipos?\s+electric", re.IGNORECASE),
]


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _safe_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        cleaned = str(value).replace(",", ".").strip()
        try:
            return float(cleaned)
        except ValueError:
            return 0.0


def _normalize(text: str) -> str:
    lowered = text.lower()
    fixes = {
        "�": "a",
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
    }
    for src, dst in fixes.items():
        lowered = lowered.replace(src, dst)
    return lowered


def _is_level_name(text: str) -> bool:
    normalized = _normalize(text)
    return any(pattern.search(normalized) for pattern in _LEVEL_PATTERNS)


def _infer_item_type(summary: str) -> str:
    normalized = _normalize(summary)
    for keywords, item_type in _ITEM_TYPE_HINTS:
        if any(keyword in normalized for keyword in keywords):
            return item_type
    return "generic_construction_item"


def _iter_budget_rows(xlsx_path: str | Path):
    workbook = load_workbook(filename=str(xlsx_path), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    # Header starts on row 3 in the known PRES format.
    for row in sheet.iter_rows(min_row=4, values_only=True):
        code = _safe_str(row[0] if len(row) > 0 else None)
        nat = _safe_str(row[1] if len(row) > 1 else None)
        unit = _safe_str(row[2] if len(row) > 2 else None)
        summary = _safe_str(row[3] if len(row) > 3 else None)
        quantity = _safe_float(row[4] if len(row) > 4 else None)
        price = _safe_float(row[5] if len(row) > 5 else None)
        amount = _safe_float(row[6] if len(row) > 6 else None)
        if not any((code, nat, unit, summary)):
            continue
        yield {
            "code": code,
            "nat": nat.lower(),
            "unit": unit,
            "summary": summary,
            "quantity": quantity,
            "price": price,
            "amount": amount,
        }


def extract_training_pairs(xlsx_path: str | Path) -> list[TrainingPair]:
    path = Path(xlsx_path)
    source_name = path.name
    pairs: list[TrainingPair] = []
    current_level = "Sin Nivel"
    current_discipline = "Sin Disciplina"

    for row in _iter_budget_rows(path):
        nat = row["nat"]
        summary = row["summary"]

        if "cap" in nat:
            if _is_level_name(summary):
                current_level = summary
                current_discipline = "General"
            else:
                current_discipline = summary or current_discipline
            continue

        if "partida" not in nat:
            continue
        if not row["code"] or not summary:
            continue

        inferred_type = _infer_item_type(summary)
        context = f"{current_level} | {current_discipline}"
        pairs.append(
            TrainingPair(
                input_item_type=inferred_type,
                input_unit=row["unit"],
                input_context=context,
                output_bc3_code=row["code"],
                output_description=summary,
                output_unit=row["unit"],
                output_quantity=row["quantity"],
                output_price=row["price"],
                source=source_name,
            )
        )

    return pairs


def extract_level_templates(xlsx_path: str | Path) -> list[LevelTemplate]:
    level_discipline_rows: dict[str, dict[str, list[TrainingPair]]] = {}

    for pair in extract_training_pairs(xlsx_path):
        level_name, _, discipline = pair.input_context.partition("|")
        level = level_name.strip() or "Sin Nivel"
        disc = discipline.strip() or "General"
        level_discipline_rows.setdefault(level, {}).setdefault(disc, []).append(pair)

    signature_groups: dict[tuple[str, str], list[str]] = {}
    payload_by_group: dict[tuple[str, str], list[TrainingPair]] = {}

    for level, disciplines in level_discipline_rows.items():
        for discipline, pairs in disciplines.items():
            signature = "|".join(
                sorted(f"{pair.output_bc3_code}:{pair.output_unit}" for pair in pairs)
            )
            group_key = (discipline, signature)
            signature_groups.setdefault(group_key, []).append(level)
            payload_by_group[group_key] = pairs

    templates: list[LevelTemplate] = []
    for index, ((discipline, signature), levels) in enumerate(
        sorted(signature_groups.items(), key=lambda item: (item[0][0], item[0][1])),
        start=1,
    ):
        pairs = payload_by_group[(discipline, signature)]
        templates.append(
            LevelTemplate(
                template_id=f"template_{index:03d}",
                levels=sorted(set(levels)),
                discipline=discipline,
                item_codes=sorted({pair.output_bc3_code for pair in pairs}),
                item_types=sorted({pair.input_item_type for pair in pairs}),
                item_count=len(pairs),
                signature=signature,
            )
        )

    return templates


# BC3 chapter codes (classifier_agent._CHAPTERS keys) -> training pair item_types to prefer
_FEW_SHOT_TYPES_BY_CHAPTER: dict[str, tuple[str, ...]] = {
    "01": ("footing", "generic_construction_item"),
    "02": ("footing", "column", "beam", "slab", "wall"),
    "03": ("wall", "wall_finish_plaster"),
    "04": ("floor_finish",),
    "05": ("door_count", "window_count"),
    "06": ("fixture_count",),
    "07": ("wet_area_fixture_count",),
    "08": ("wall_finish_paint",),
    "09": ("generic_construction_item",),
}

# First word of chapter title (lowercase) -> item types when chapter_code is missing
_FEW_SHOT_TYPES_BY_TITLE_WORD: dict[str, tuple[str, ...]] = {
    "movimiento": ("footing", "generic_construction_item"),
    "hormigon": ("footing", "column", "beam", "slab", "wall"),
    "muros": ("wall", "wall_finish_plaster"),
    "pisos": ("floor_finish",),
    "puertas": ("door_count", "window_count"),
    "ventanas": ("window_count",),
    "instalaciones": ("fixture_count", "wet_area_fixture_count"),
    "sanitarias": ("wet_area_fixture_count",),
    "pintura": ("wall_finish_paint", "wall_finish_plaster"),
    "gastos": ("generic_construction_item",),
    "electrico": ("fixture_count",),
    "sanitario": ("wet_area_fixture_count",),
}


def generate_few_shot_examples(
    training_pairs: list[TrainingPair],
    category: str,
    *,
    chapter_code: str | None = None,
    max_examples: int = 10,
) -> str:
    """Few-shot lines use PRES reference rows (description + item_type); BC3 codes in examples are illustrative only."""
    if not training_pairs:
        return ""

    normalized_category = _normalize(category)
    allowed_types: tuple[str, ...] = ()
    if chapter_code and chapter_code in _FEW_SHOT_TYPES_BY_CHAPTER:
        allowed_types = _FEW_SHOT_TYPES_BY_CHAPTER[chapter_code]
    elif normalized_category in _FEW_SHOT_TYPES_BY_TITLE_WORD:
        allowed_types = _FEW_SHOT_TYPES_BY_TITLE_WORD[normalized_category]
    else:
        legacy_map: dict[str, tuple[str, ...]] = {
            "muros": ("wall", "wall_finish_plaster"),
            "pisos": ("floor_finish",),
            "puertas": ("door_count",),
            "ventanas": ("window_count",),
            "hormigon": ("footing", "column", "beam", "slab"),
            "electrico": ("fixture_count",),
            "sanitario": ("wet_area_fixture_count",),
        }
        allowed_types = legacy_map.get(normalized_category, ())

    filtered = [
        pair
        for pair in training_pairs
        if (
            (allowed_types and pair.input_item_type in allowed_types)
            or normalized_category in _normalize(pair.output_description)
            or normalized_category in _normalize(pair.input_context)
        )
    ]
    if not filtered:
        filtered = training_pairs

    selected = filtered[: max(1, min(max_examples, 12))]
    header = "EJEMPLOS REALES (PRES / BC3 — few-shot):"
    if chapter_code:
        header += f" capítulo {chapter_code}."
    lines = [header]
    for pair in selected:
        lines.append(
            (
                f"- input: item_type={pair.input_item_type}, unit={pair.input_unit}, "
                f"context={pair.input_context}\n"
                f"  output: bc3_code={pair.output_bc3_code}, desc={pair.output_description}, "
                f"unit={pair.output_unit}, qty={pair.output_quantity:.2f}, price={pair.output_price:.2f}"
            )
        )
    lines.append(
        "Regla: los códigos anteriores son del proyecto de referencia; en esta tarea solo "
        "puedes usar códigos que aparezcan en el catálogo BC3 proporcionado abajo."
    )
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Multi-source corpus loader
# ---------------------------------------------------------------------------

def load_training_corpus(
    paths: list[str | Path],
    *,
    deduplicate: bool = True,
) -> list[TrainingPair]:
    """
    Load and merge training pairs from multiple XLSX budget files.

    Parameters
    ----------
    paths:
        List of paths to XLSX files (e.g., multiple project budgets from Dropbox).
        Non-existent paths are skipped with a warning.
    deduplicate:
        When True (default), removes pairs whose (code, description) combination
        is already present from a previous file in the list.  This avoids
        over-representing projects that appear in multiple formats.

    Returns
    -------
    Merged list of ``TrainingPair`` instances, preserving insertion order.
    """
    merged: list[TrainingPair] = []
    seen_keys: set[tuple[str, str]] = set()

    for raw_path in paths:
        path = Path(raw_path)
        if not path.exists():
            logger.warning("Training file not found, skipping: %s", path)
            continue
        try:
            pairs = extract_training_pairs(path)
            added = 0
            for pair in pairs:
                key = (pair.output_bc3_code.strip(), pair.output_description.strip())
                if deduplicate and key in seen_keys:
                    continue
                seen_keys.add(key)
                merged.append(pair)
                added += 1
            logger.info(
                "Loaded %d training pairs (%d new) from %s",
                len(pairs), added, path.name,
            )
        except Exception:
            logger.warning("Failed to load training pairs from %s", path, exc_info=True)

    logger.info("Training corpus total: %d pairs from %d files", len(merged), len(paths))
    return merged


def load_training_corpus_from_folder(
    folder: str | Path,
    *,
    pattern: str = "*.xlsx",
    deduplicate: bool = True,
) -> list[TrainingPair]:
    """
    Load training pairs from all XLSX files found in ``folder``.

    Parameters
    ----------
    folder:
        Directory to scan (non-recursive by default via ``pattern``).
        Use ``**/*.xlsx`` for a recursive scan.
    pattern:
        Glob pattern used to locate files inside ``folder``.
    deduplicate:
        Forwarded to ``load_training_corpus``.

    Returns
    -------
    Merged list of ``TrainingPair`` instances.
    """
    folder_path = Path(folder)
    if not folder_path.is_dir():
        raise NotADirectoryError(f"Training folder not found: {folder_path}")

    # Use rglob for patterns that include ** (recursive), glob otherwise.
    if "**" in pattern:
        sub_pattern = pattern.lstrip("*").lstrip("/")
        xlsx_paths = sorted(folder_path.rglob(sub_pattern))
    else:
        xlsx_paths = sorted(folder_path.glob(pattern))
    if not xlsx_paths:
        logger.warning("No XLSX files found in %s (pattern=%r)", folder_path, pattern)
        return []

    logger.info(
        "Scanning %s for training data (%d files matched %r)",
        folder_path, len(xlsx_paths), pattern,
    )
    return load_training_corpus(xlsx_paths, deduplicate=deduplicate)


def export_corpus_jsonl(
    training_pairs: list[TrainingPair],
    output_path: str | Path,
) -> Path:
    """
    Export training pairs to a JSONL file suitable for fine-tuning or offline analysis.

    Each line is a JSON object with the fine-tuning message format used by
    ``FeedbackStore.export_for_fine_tuning``.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as handle:
        for pair in training_pairs:
            record = {
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "Eres un asistente de presupuesto de construcción dominicano. "
                            "Dado un tipo de elemento y contexto de nivel/disciplina, "
                            "devuelve el codigo BC3 correcto, descripcion, unidad y precio."
                        ),
                    },
                    {
                        "role": "user",
                        "content": (
                            f"item_type={pair.input_item_type}\n"
                            f"unit={pair.input_unit}\n"
                            f"context={pair.input_context}"
                        ),
                    },
                    {
                        "role": "assistant",
                        "content": json.dumps(
                            {
                                "bc3_code": pair.output_bc3_code,
                                "description": pair.output_description,
                                "unit": pair.output_unit,
                                "quantity": pair.output_quantity,
                                "price": pair.output_price,
                            },
                            ensure_ascii=False,
                        ),
                    },
                ],
                "metadata": {"source": pair.source},
            }
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")
    logger.info("Exported %d training pairs to %s", len(training_pairs), out)
    return out
