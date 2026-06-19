"""
Multi-discipline budget consolidator.

Merges per-discipline budget outputs into a single multi-sheet Excel workbook
with a summary sheet referencing each discipline's subtotal.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.schemas import BudgetRow, ProjectContext

logger = logging.getLogger("dupla.consolidator")

# Reuse styling constants from the single-sheet exporter
HEADERS = ("Codigo", "Nat", "Ud", "Resumen", "CanPres", "PrPres", "ImpPres")
THIN_SIDE = Side(style="thin", color="BFBFBF")
ALL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
CHAPTER_FILL = PatternFill("solid", fgColor="FFF2CC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
SUMMARY_FILL = PatternFill("solid", fgColor="4472C4")
SUMMARY_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="2F5496")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=12)


DISCIPLINE_ORDER = ["arquitectura", "estructura", "electrico", "sanitario"]

_INVALID_SHEET_CHARS = frozenset("\\/*?:[]")


def _excel_sheet_title(name: str, *, max_len: int = 31) -> str:
    """Excel sheet titles cannot contain \\ / * ? : [ ]."""
    cleaned = "".join(" " if ch in _INVALID_SHEET_CHARS else ch for ch in name)
    cleaned = " ".join(cleaned.split()).strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:max_len]


def _unique_sheet_title(name: str, existing: set[str]) -> str:
    base = _excel_sheet_title(name)
    candidate = base
    suffix_n = 2
    while candidate in existing:
        suffix = f" {suffix_n}"
        candidate = _excel_sheet_title(name, max_len=31 - len(suffix)) + suffix
        suffix_n += 1
    existing.add(candidate)
    return candidate


def _coerce_row(row: BudgetRow | Mapping[str, object]) -> BudgetRow:
    if isinstance(row, BudgetRow):
        return row
    payload = dict(row)
    return BudgetRow(
        row_type=str(payload.get("row_type", "line")),
        code=str(payload.get("code", "")),
        nat=str(payload.get("nat", "")),
        unit=str(payload.get("unit", "")),
        summary=str(payload.get("summary", "")),
        quantity=payload.get("quantity"),
        unit_price=payload.get("unit_price"),
        amount=payload.get("amount"),
        chapter_id=payload.get("chapter_id"),
        parent_chapter_id=payload.get("parent_chapter_id"),
        level=int(payload.get("level", 0) or 0),
        takeoff_key=payload.get("takeoff_key"),
        source_refs=list(payload.get("source_refs", [])),
        assumptions=list(payload.get("assumptions", [])),
        metadata=dict(payload.get("metadata", {})),
        excel_row=payload.get("excel_row"),
    )


def _write_discipline_sheet(
    workbook: Workbook,
    sheet_name: str,
    discipline_display: str,
    project_title: str,
    rows: Iterable[BudgetRow | Mapping[str, object]],
) -> int | None:
    """Write a single discipline sheet. Returns the excel row of the grand subtotal (last subtotal row), or None."""
    ws = workbook.create_sheet(title=sheet_name)

    ws["A1"] = project_title
    ws["A2"] = discipline_display
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"].font = Font(size=12, bold=True)

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = ALL_BORDER

    coerced = [_coerce_row(r) for r in rows]
    last_subtotal_row: int | None = None

    for row in coerced:
        target_row = row.excel_row or 4
        values = (
            row.code, 
            row.nat, 
            row.unit, 
            row.summary, 
            row.metadata.get("excel_quantity_formula", row.quantity), 
            row.metadata.get("excel_unit_price_formula", row.unit_price), 
            row.metadata.get("excel_amount_formula", row.amount)
        )
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=target_row, column=col_idx)
            cell.value = value
            cell.border = ALL_BORDER
            if col_idx >= 5:
                cell.number_format = "#,##0.00"

        row_fill = None
        row_font = Font(bold=False)
        if row.row_type == "chapter":
            row_fill = CHAPTER_FILL
            row_font = Font(bold=True)
        elif row.row_type == "subtotal":
            row_fill = SUBTOTAL_FILL
            row_font = Font(bold=True)
            last_subtotal_row = target_row

        for col_idx in range(1, 8):
            cell = ws.cell(row=target_row, column=col_idx)
            cell.font = row_font
            cell.alignment = Alignment(
                horizontal="left" if col_idx <= 4 else "right",
                vertical="center",
            )
            if row_fill is not None:
                cell.fill = row_fill

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16

    return last_subtotal_row


def _write_summary_sheet(
    workbook: Workbook,
    project_title: str,
    discipline_refs: list[tuple[str, str, int | None]],
) -> None:
    """Write the RESUMEN GENERAL summary sheet as the first sheet.

    Args:
        discipline_refs: List of (discipline_display, sheet_name, subtotal_row).
            subtotal_row may be None if the discipline had no rows.
    """
    ws = workbook.create_sheet(title="RESUMEN GENERAL", index=0)

    ws["A1"] = project_title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = "RESUMEN GENERAL POR DISCIPLINA"
    ws["A2"].font = Font(size=12, bold=True)
    ws["A3"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = Font(size=9, italic=True, color="666666")

    summary_headers = ("Disciplina", "Subtotal RD$", "% del Total", "Observaciones")
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = SUMMARY_FONT
        cell.fill = SUMMARY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = ALL_BORDER

    data_start = 6
    for i, (display_name, sheet_name, subtotal_row) in enumerate(discipline_refs):
        row_num = data_start + i
        ws.cell(row=row_num, column=1, value=display_name).border = ALL_BORDER

        amount_cell = ws.cell(row=row_num, column=2)
        if subtotal_row is not None:
            amount_cell.value = f"='{sheet_name}'!G{subtotal_row}"
        else:
            amount_cell.value = 0
        amount_cell.number_format = "#,##0.00"
        amount_cell.border = ALL_BORDER

        pct_cell = ws.cell(row=row_num, column=3)
        pct_cell.border = ALL_BORDER

        obs_cell = ws.cell(row=row_num, column=4)
        if subtotal_row is None:
            obs_cell.value = "Sin partidas"
        obs_cell.border = ALL_BORDER

    total_row = data_start + len(discipline_refs)
    ws.cell(row=total_row, column=1, value="TOTAL GENERAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.cell(row=total_row, column=1).border = ALL_BORDER

    total_cell = ws.cell(row=total_row, column=2)
    sum_range = f"B{data_start}:B{total_row - 1}"
    total_cell.value = f"=SUM({sum_range})"
    total_cell.font = TOTAL_FONT
    total_cell.fill = TOTAL_FILL
    total_cell.number_format = "#,##0.00"
    total_cell.border = ALL_BORDER

    for col_idx in range(3, 5):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = TOTAL_FILL
        cell.border = ALL_BORDER

    # Percentage formulas
    for i in range(len(discipline_refs)):
        row_num = data_start + i
        pct_cell = ws.cell(row=row_num, column=3)
        pct_cell.value = f"=IF(B{total_row}=0,0,B{row_num}/B{total_row})"
        pct_cell.number_format = "0.0%"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 25


def _save_workbook(workbook: Workbook, output: Path, *, max_fallback: int = 20) -> Path:
    try:
        workbook.save(output)
        return output
    except PermissionError as exc:
        last_error = exc
        for attempt in range(1, max_fallback + 1):
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            suffix = "" if attempt == 1 else f"_{attempt:02d}"
            fallback = output.with_name(f"{output.stem}_{ts}{suffix}{output.suffix}")
            try:
                workbook.save(fallback)
                return fallback
            except PermissionError as fe:
                last_error = fe
    raise PermissionError(f"Cannot save to '{output}' or fallbacks.") from last_error


_DISPLAY_NAMES: dict[str, str] = {
    "arquitectura": "ARQUITECTURA",
    "estructura": "ESTRUCTURA",
    "electrico": "EL\u00c9CTRICO",
    "sanitario": "SANITARIO / PLOMER\u00cdA",
}


def consolidate_budgets(
    discipline_budgets: dict[str, dict[str, Any]],
    output_path: str | Path,
    project_name: str,
) -> Path:
    """Write multi-sheet Excel workbook with per-discipline sheets + summary.

    Args:
        discipline_budgets: Mapping of discipline_id to the output of
            ``compose_budget()`` for that discipline (must contain ``rows``).
        output_path: Destination ``.xlsx`` path.
        project_name: Project title for headers.

    Returns:
        Path to the written workbook.
    """
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    # Remove the default sheet created by openpyxl
    if workbook.active is not None:
        workbook.remove(workbook.active)

    ordered_ids = [d for d in DISCIPLINE_ORDER if d in discipline_budgets]
    ordered_ids.extend(d for d in sorted(discipline_budgets) if d not in ordered_ids)

    discipline_refs: list[tuple[str, str, int | None]] = []
    used_sheet_titles: set[str] = {"RESUMEN GENERAL"}

    for disc_id in ordered_ids:
        budget = discipline_budgets[disc_id]
        rows = budget.get("rows", [])
        display_name = _DISPLAY_NAMES.get(disc_id, disc_id.upper())
        sheet_name = _unique_sheet_title(display_name, used_sheet_titles)

        subtotal_row = _write_discipline_sheet(
            workbook, sheet_name, display_name, project_name, rows,
        )
        discipline_refs.append((display_name, sheet_name, subtotal_row))
        logger.info(
            "Discipline '%s': %d rows, subtotal row %s",
            disc_id, len(rows), subtotal_row,
        )

    _write_summary_sheet(workbook, project_name, discipline_refs)

    saved_path = _save_workbook(workbook, output)
    logger.info("Consolidated workbook saved: %s (%d disciplines)", saved_path, len(ordered_ids))
    return saved_path
