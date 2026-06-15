"""
Filtra un libro PRES (formato Presto habitual) a líneas de alcance estructural.

Sirve para comparar el presupuesto derivado de un DWG/plano principalmente estructural
contra una línea base PREC acotada (sin albañilería, instalaciones, acabados, etc.).

La heurística combina etiquetas de disciplina (mismas ideas que compare_budget) con
palabras clave en el resumen. Es revisable y extensible.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from compare_budget import _discipline_tags

# Subconjunto de disciplinas que compare_budget etiqueta como “estructura / cimentación”.
STRUCTURAL_DISCIPLINE_TAGS: frozenset[str] = frozenset(
    {
        "movimiento_tierra",
        "hormigon_armado",
        "acero_refuerzo",
    }
)

# Capítulos típicos en PRES (texto en columna resumen / Nat).
_STRUCT_CHAPTER_HINTS: tuple[str, ...] = (
    "estructur",
    "hormigon",
    "hormigón",
    "ciment",
    "zapata",
    "losas",
    "losa",
    "viga",
    "column",
    "columna",
    "encofr",
    "cimbra",
    "acero",
    "concreto",
    "movimiento de tierra",
    "excavac",
    "relleno",
    "compactac",
    "demolic",
    "platea",
    "radier",
    "fundacion",
    "fundación",
)

# Partidas: refuerzo por texto si el clasificador de disciplinas no las etiqueta.
_STRUCT_PARTIDA_HINTS: tuple[str, ...] = (
    "hormigon",
    "hormigón",
    "concreto",
    "zapata",
    "platea",
    "radier",
    "cimient",
    "encofr",
    "cimbra",
    "viga",
    "losa",
    "columna",
    "column",
    "poste",
    "acero",
    "armad",
    "estrib",
    "refuerz",
    "pretensad",
    "estructur",
    "excavac",
    "relleno",
    "compactac",
    "bote de material",
    "movimiento de tierra",
)

# Excluir acabados / instalaciones aunque aparezca una palabra ambigua.
_STRUCT_EXCLUDE_HINTS: tuple[str, ...] = (
    "panete",
    "pañete",
    "fraguache",
    "ceramic",
    "cerámica",
    "porcelan",
    "pintura",
    "electr",
    "eléctric",
    "sanitar",
    "plomer",
    "inodor",
    "lavamanos",
    "puerta",
    "ventana",
    "closet",
    "gabinete",
    "zocalo",
    "zócalo",
    "terminacion",
    "terminación",
    "acabad",
)


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize(text: str) -> str:
    lowered = text.lower()
    for src, dst in (
        ("á", "a"),
        ("é", "e"),
        ("í", "i"),
        ("ó", "o"),
        ("ú", "u"),
        ("ñ", "n"),
    ):
        lowered = lowered.replace(src, dst)
    return lowered


def row_is_structural_partida(code: str, nat: str, summary: str) -> bool:
    if "partida" not in _normalize(nat):
        return False
    blob = f"{_normalize(code)} {_normalize(summary)}"
    if any(h in blob for h in _STRUCT_EXCLUDE_HINTS):
        if not any(h in blob for h in ("hormigon", "hormigón", "concreto", "zapata", "viga", "losa", "column", "acero")):
            return False
    tags = _discipline_tags(summary)
    if tags & STRUCTURAL_DISCIPLINE_TAGS:
        return True
    if any(h in blob for h in _STRUCT_PARTIDA_HINTS):
        return True
    return False


def row_is_structural_chapter(nat: str, summary: str) -> bool:
    if "cap" not in _normalize(nat):
        return False
    blob = _normalize(summary)
    return any(h in blob for h in _STRUCT_CHAPTER_HINTS)


def _row_tuple(sheet: Any, row: int) -> tuple[Any, ...]:
    return tuple(sheet.cell(row=row, column=c).value for c in range(1, 8))


def filter_pres_workbook_structural(src: Path, dest: Path) -> dict[str, int]:
    """
    Escribe un nuevo .xlsx con filas 1-3 idénticas y solo capítulos/partidas estructurales.

    Returns:
        Contadores útiles para logging (input_rows, kept_rows, partidas_kept).
    """
    wb = load_workbook(filename=str(src), data_only=False)
    ws = wb[wb.sheetnames[0]]
    max_row = ws.max_row or 4

    body: list[tuple[int, tuple[Any, ...]]] = []
    for r in range(4, max_row + 1):
        vals = _row_tuple(ws, r)
        code = _safe_str(vals[0] if len(vals) > 0 else None)
        nat = _safe_str(vals[1] if len(vals) > 1 else None)
        unit = _safe_str(vals[2] if len(vals) > 2 else None)
        summary = _safe_str(vals[3] if len(vals) > 3 else None)
        if not any((code, nat, unit, summary)):
            continue
        body.append((r, vals))

    kept: list[tuple[Any, ...]] = []
    pending_chapters: list[tuple[Any, ...]] = []
    partidas_kept = 0

    for _r, vals in body:
        code = _safe_str(vals[0] if len(vals) > 0 else None)
        nat = _safe_str(vals[1] if len(vals) > 1 else None)
        summary = _safe_str(vals[3] if len(vals) > 3 else None)

        if "cap" in _normalize(nat):
            if row_is_structural_chapter(nat, summary):
                pending_chapters.append(vals)
            else:
                pending_chapters.clear()
            continue

        if "partida" in _normalize(nat):
            if row_is_structural_partida(code, nat, summary):
                kept.extend(pending_chapters)
                pending_chapters.clear()
                kept.append(vals)
                partidas_kept += 1
            else:
                pending_chapters.clear()
            continue

        pending_chapters.clear()

    out = Workbook()
    ows = out.active
    ows.title = ws.title
    for r in range(1, 4):
        for c in range(1, 8):
            ows.cell(row=r, column=c, value=ws.cell(row=r, column=c).value)

    out_row = 4
    for vals in kept:
        for c, value in enumerate(vals, start=1):
            ows.cell(row=out_row, column=c, value=value)
        out_row += 1

    dest.parent.mkdir(parents=True, exist_ok=True)
    out.save(str(dest))

    return {
        "input_body_rows": len(body),
        "kept_rows": len(kept),
        "partidas_kept": partidas_kept,
    }
