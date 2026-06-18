"""
Excel export for composed Dupla budget rows.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.schemas import BudgetRow, ProjectContext

HEADERS = (
    "Código",
    "Nat",
    "Ud",
    "Resumen",
    "CanPres",
    "PrPres",
    "ImpPres",
    "Fuente Cantidad",
    "Fuente Precio",
    "BC3 Origen",
    "Método de Precio",
    "Revisión",
    "Confianza",
    "Referencias",
    "Supuestos",
    "Código APU",
    "Desglose APU",
)
REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")
PENDING_FILL = PatternFill("solid", fgColor="FFFF00")
THIN_SIDE = Side(style="thin", color="BFBFBF")
ALL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
CHAPTER_FILL = PatternFill("solid", fgColor="FFF2CC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")


def _format_apu_components(metadata: Mapping[str, Any]) -> str:
    parts: list[str] = []
    for comp in metadata.get("apu_components") or []:
        if not isinstance(comp, dict):
            continue
        desc = str(comp.get("description") or "")[:40]
        qty = comp.get("quantity")
        unit = comp.get("unit") or ""
        price = comp.get("unit_price")
        if desc:
            parts.append(f"{desc} {qty}{unit}@{price}")
    return "; ".join(parts[:6])


def _coerce_row(row: BudgetRow | Mapping[str, object]) -> BudgetRow:
    if isinstance(row, BudgetRow):
        return row
    payload = dict(row)
    return BudgetRow(
        row_type=str(payload.get("row_type", "line")),
        code=str(payload.get("code", "")),
        nat=str(payload.get("nat", "")),
        unit=str(payload.get("unit", "")),
        summary=str(payload.get("summary", "")),
        quantity=payload.get("quantity"),
        unit_price=payload.get("unit_price"),
        amount=payload.get("amount"),
        chapter_id=payload.get("chapter_id"),
        parent_chapter_id=payload.get("parent_chapter_id"),
        level=int(payload.get("level", 0) or 0),
        takeoff_key=payload.get("takeoff_key"),
        source_refs=list(payload.get("source_refs", [])),
        assumptions=list(payload.get("assumptions", [])),
        metadata=dict(payload.get("metadata", {})),
        excel_row=payload.get("excel_row"),
    )


def _write_value(cell, value: object) -> None:
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        return
    if value is None:
        cell.value = None


def _fallback_output_path(output: Path, attempt: int) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = "" if attempt == 1 else f"_{attempt:02d}"
    return output.with_name(f"{output.stem}_{timestamp}{suffix}{output.suffix}")


def _save_workbook(workbook: Workbook, output: Path, *, max_fallback_attempts: int = 20) -> Path:
    try:
        workbook.save(output)
        return output
    except PermissionError as exc:
        last_error = exc
        for attempt in range(1, max_fallback_attempts + 1):
            fallback_output = _fallback_output_path(output, attempt)
            try:
                workbook.save(fallback_output)
                return fallback_output
            except PermissionError as fallback_exc:
                last_error = fallback_exc

    raise PermissionError(
        f"Could not save workbook to '{output}' or any fallback filename in '{output.parent}'. "
        "Close the workbook if it is open, or verify write permissions for the directory."
    ) from last_error


def _append_quality_sheet(workbook: Workbook, quality_report: Mapping[str, Any], *, sheet_name: str = "Quality_Report") -> None:
    worksheet = workbook.create_sheet(title=sheet_name)
    headers = (
        "status",
        "code",
        "discipline",
        "element_id",
        "level_id",
        "unit_id",
        "space_id",
        "confidence",
        "message",
        "evidence",
        "suggested_action",
    )
    for idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = ALL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    issues = list(quality_report.get("issues", []))
    for row_index, issue in enumerate(issues, start=2):
        evidence = issue.get("evidence_refs") or []
        values = (
            issue.get("status", ""),
            issue.get("code", ""),
            issue.get("discipline", ""),
            issue.get("element_id", ""),
            issue.get("level_id", ""),
            issue.get("unit_id", ""),
            issue.get("space_id", ""),
            issue.get("confidence_score", ""),
            issue.get("message", ""),
            ", ".join(str(item) for item in evidence if item),
            issue.get("suggested_action", ""),
        )
        for col_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=col_index, value=value)
            cell.border = ALL_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 12
    worksheet.column_dimensions["B"].width = 22
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["E"].width = 14
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 14
    worksheet.column_dimensions["H"].width = 12
    worksheet.column_dimensions["I"].width = 58
    worksheet.column_dimensions["J"].width = 46
    worksheet.column_dimensions["K"].width = 46


def export_budget_workbook(
    context: ProjectContext,
    rows: Iterable[BudgetRow | Mapping[str, object]],
    output_path: str | Path,
    *,
    sheet_name: str = "Presupuesto",
    quality_report: Mapping[str, Any] | None = None,
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    project_title = context.project_name or context.project_id or "DUPLA"
    worksheet["A1"] = project_title
    worksheet["A2"] = "Presupuesto"

    worksheet["A1"].font = Font(size=14, bold=True)
    worksheet["A2"].font = Font(size=12, bold=True)

    for column_index, header in enumerate(HEADERS, start=1):
        cell = worksheet.cell(row=3, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = ALL_BORDER

    coerced_rows = [_coerce_row(row) for row in rows]
    for row in coerced_rows:
        target_row = row.excel_row or 4
        price_source = ""
        quantity_source = ""
        bc3_origin = ""
        candidate_source = ""
        requiere_revision_text = ""
        confidence_value: float | str = ""
        source_refs_text = ""
        assumptions_text = ""
        apu_code = ""
        apu_desglose = ""
        needs_review = False
        if row.row_type == "line":
            price_source = str(row.metadata.get("price_source") or "")
            quantity_source = str(row.metadata.get("quantity_source_display") or "")
            bc3_origin = str(row.metadata.get("bc3_origin") or "")
            candidate_source = str(row.metadata.get("candidate_source") or "")
            needs_review = bool(row.metadata.get("requiere_revision"))
            requiere_revision_text = "Sí" if needs_review else "No"
            raw_conf = row.metadata.get("confidence")
            if raw_conf is not None:
                try:
                    confidence_value = round(float(raw_conf), 2)
                except (TypeError, ValueError):
                    confidence_value = ""
            source_refs_text = "; ".join(str(r) for r in (row.source_refs or [])[:5])
            assumptions_text = "; ".join(str(a) for a in (row.assumptions or [])[:3])
            apu_code = str(row.metadata.get("apu_code") or row.metadata.get("candidate_code") or "")
            apu_desglose = _format_apu_components(row.metadata)

        values = (
            row.code,
            row.nat,
            row.unit,
            row.summary,
            row.quantity,
            row.unit_price,
            row.amount,
            quantity_source,
            price_source,
            bc3_origin,
            candidate_source,
            requiere_revision_text,
            confidence_value,
            source_refs_text,
            assumptions_text,
            apu_code,
            apu_desglose,
        )
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=target_row, column=column_index)
            _write_value(cell, value)
            cell.border = ALL_BORDER
            if 5 <= column_index <= 7:
                cell.number_format = '#,##0.00'

        row_fill = None
        row_font = Font(bold=False)
        if row.row_type == "chapter":
            row_fill = CHAPTER_FILL
            row_font = Font(bold=True)
        elif row.row_type == "subtotal":
            row_fill = SUBTOTAL_FILL
            row_font = Font(bold=True)
        elif row.row_type == "line" and needs_review:
            row_fill = REVIEW_FILL

        for column_index in range(1, len(HEADERS) + 1):
            cell = worksheet.cell(row=target_row, column=column_index)
            cell.font = row_font
            cell.alignment = Alignment(
                horizontal="left" if column_index <= 4 or column_index >= 8 else "right",
                vertical="center",
            )
            if row_fill is not None:
                cell.fill = row_fill

    worksheet.freeze_panes = "A4"
    worksheet.sheet_view.showGridLines = True
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 16
    worksheet.column_dimensions["C"].width = 10
    worksheet.column_dimensions["D"].width = 60
    worksheet.column_dimensions["E"].width = 14
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 16
    worksheet.column_dimensions["H"].width = 32
    worksheet.column_dimensions["I"].width = 28
    worksheet.column_dimensions["J"].width = 22
    worksheet.column_dimensions["K"].width = 22
    worksheet.column_dimensions["L"].width = 12
    worksheet.column_dimensions["M"].width = 12
    worksheet.column_dimensions["N"].width = 36
    worksheet.column_dimensions["O"].width = 36
    worksheet.column_dimensions["P"].width = 18
    worksheet.column_dimensions["Q"].width = 48

    if quality_report:
        _append_quality_sheet(workbook, quality_report)

    _append_pendientes_sheet(workbook, coerced_rows)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    return _save_workbook(workbook, output)


def _append_pendientes_sheet(
    workbook: Workbook,
    rows: list[Any],
    *,
    sheet_name: str = "PENDIENTES",
) -> None:
    """
    Creates a separate sheet listing all budget line rows where unit_price is None.
    These are items that ConstruCosto could not price and require manual review.
    """
    pending = [
        row for row in rows
        if isinstance(row, object)
        and getattr(row, "row_type", None) == "line"
        and getattr(row, "unit_price", 1) is None
    ]
    if not pending:
        return

    worksheet = workbook.create_sheet(title=sheet_name)
    headers = ("Código", "Ud", "Resumen", "CanPres", "Takeoff Key", "Disciplina")
    for idx, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.border = ALL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(pending, start=2):
        discipline = str(row.metadata.get("source_discipline") or "") if hasattr(row, "metadata") else ""
        takeoff_key = str(getattr(row, "takeoff_key", "") or "")
        values = (
            getattr(row, "code", ""),
            getattr(row, "unit", ""),
            getattr(row, "summary", ""),
            getattr(row, "quantity", None),
            takeoff_key,
            discipline,
        )
        for col_idx, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=col_idx, value=value)
            cell.fill = PENDING_FILL
            cell.border = ALL_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 10
    worksheet.column_dimensions["C"].width = 60
    worksheet.column_dimensions["D"].width = 14
    worksheet.column_dimensions["E"].width = 42
    worksheet.column_dimensions["F"].width = 18
