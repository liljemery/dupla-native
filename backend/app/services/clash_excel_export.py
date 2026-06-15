"""Excel export for clash corrida and final workflow reports."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from app.domain.clash_workflow_enums import decision_label, status_label
from app.domain.clash_workflow_enums import ClashStatus, ReviewerDecision
from app.models.project_clash_item import ProjectClashItem
from app.services.clash_reports.data import ReportBundle


def _header_row(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)


def build_corrida_technical_excel(bundle: ReportBundle) -> bytes:
    """Technical corrida spreadsheet from artifact bundle (pre-review)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidencias"
    headers = [
        "Código",
        "Severidad",
        "Confianza",
        "Nivel",
        "DWG A",
        "DWG B",
        "Capa A",
        "Capa B",
        "Disciplina A",
        "Disciplina B",
        "Área m²",
        "Prof. Z mm",
        "Tipo",
        "Qué revisar",
    ]
    _header_row(ws, headers)
    for row_idx, inc in enumerate(bundle.incidents, start=2):
        ws.cell(row=row_idx, column=1, value=inc.human_code)
        ws.cell(row=row_idx, column=2, value=inc.severity)
        ws.cell(row=row_idx, column=3, value=inc.confidence)
        ws.cell(row=row_idx, column=4, value=inc.level_id)
        ws.cell(row=row_idx, column=5, value=inc.file_a_alias)
        ws.cell(row=row_idx, column=6, value=inc.file_b_alias)
        ws.cell(row=row_idx, column=7, value=inc.layer_a)
        ws.cell(row=row_idx, column=8, value=inc.layer_b)
        ws.cell(row=row_idx, column=9, value=inc.discipline_a)
        ws.cell(row=row_idx, column=10, value=inc.discipline_b)
        ws.cell(row=row_idx, column=11, value=inc.area_m2_text)
        ws.cell(row=row_idx, column=12, value=inc.z_depth_text)
        ws.cell(row=row_idx, column=13, value=inc.clash_type)
        ws.cell(row=row_idx, column=14, value=inc.what_to_check)

    meta = wb.create_sheet("Metadatos")
    for i, (k, v) in enumerate(bundle.meta.items(), start=1):
        meta.cell(row=i, column=1, value=str(k))
        meta.cell(row=i, column=2, value=str(v))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_final_technical_excel(
    *,
    meta: dict[str, Any],
    items: list[ProjectClashItem],
) -> bytes:
    """Final technical export including reviewer decisions (or defaults)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clashes y decisiones"
    headers = [
        "Código",
        "Prioridad",
        "Severidad",
        "Estado",
        "Decisión revisor",
        "DWG A",
        "DWG B",
        "Nivel",
        "Capas",
        "Responsable",
        "Observación",
        "Acción recomendada",
        "Área mm²",
        "Solape Z mm",
    ]
    _header_row(ws, headers)
    for row_idx, item in enumerate(items, start=2):
        try:
            st = status_label(ClashStatus(item.status))
        except ValueError:
            st = item.status
        dec_text = "—"
        if item.reviewer_decision:
            try:
                dec_text = decision_label(ReviewerDecision(item.reviewer_decision))
            except ValueError:
                dec_text = item.reviewer_decision
        layers = " / ".join(x for x in (item.layer_a, item.layer_b) if x)
        ws.cell(row=row_idx, column=1, value=item.clash_code)
        ws.cell(row=row_idx, column=2, value=item.priority)
        ws.cell(row=row_idx, column=3, value=item.severity)
        ws.cell(row=row_idx, column=4, value=st)
        ws.cell(row=row_idx, column=5, value=dec_text)
        ws.cell(row=row_idx, column=6, value=item.dwg_a)
        ws.cell(row=row_idx, column=7, value=item.dwg_b)
        ws.cell(row=row_idx, column=8, value=item.level_id)
        ws.cell(row=row_idx, column=9, value=layers)
        ws.cell(row=row_idx, column=10, value=item.assigned_to)
        ws.cell(row=row_idx, column=11, value=item.observation)
        ws.cell(row=row_idx, column=12, value=item.recommended_action)
        ws.cell(row=row_idx, column=13, value=item.area_mm2)
        ws.cell(row=row_idx, column=14, value=item.overlap_depth_mm)

    meta_ws = wb.create_sheet("Metadatos")
    for i, (k, v) in enumerate(meta.items(), start=1):
        meta_ws.cell(row=i, column=1, value=str(k))
        meta_ws.cell(row=i, column=2, value=str(v))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
