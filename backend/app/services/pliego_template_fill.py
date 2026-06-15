"""Rellena la plantilla oficial GA-FO-01 Pliego de Condiciones - Arquitectura (1:1 en layout)."""

from __future__ import annotations

import re
import unicodedata
from copy import copy
from io import BytesIO
from pathlib import Path
from typing import Any, Optional, Tuple

from openpyxl.cell import MergedCell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

PLIEGO_TEMPLATE_FILENAMES: tuple[str, ...] = (
    "GA-FO-01-(06-2025)-V02- Pliego de Condiciones - Arquitectura.xlsx",
    "GA-FO-01-pliego.xlsx",
)

RESUMEN_SHEET_NAME = "RESUMEN"
# Hoja RESUMEN GA-FO-01: columna A = N.º partida; D = Estado; F = Observaciones
RESUMEN_COL_ESTADO = 4
RESUMEN_COL_OBSERVACIONES = 6

GA_FO_01_ESTADO_LABELS: dict[str, str] = {
    "PENDIENTE": "Pendiente",
    "COMPLETO": "Completo",
    "INCOMPLETO": "Incompleto",
    "EN_REVISION": "En revisión",
    "NO_APLICA": "No aplica",
}

# (campo_payload, palabras clave en cabecera normalizada)
HEADER_KEYWORDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("partida", ("partida", "item", "item.", "n°", "nº", "no.", "cod", "codigo", "código")),
    ("descripcion", ("descripcion", "detalle", "concepto")),
    ("unidad", ("unidad", "und", "ud.", "ud", "ume")),
    ("cantidad", ("cantidad", "cant")),
    ("precio_unitario", ("precio unitario", "p. unit", "p.unit", "v.unit", "valor unitario", "unitario", "v. unit")),
    ("subtotal", ("subtotal", "importe", "parcial", "valor total")),
    ("notas", ("notas", "observaciones", "obs")),
    ("capitulo", ("capitulo", "capítulo", "cap")),
    ("grupo", ("grupo", "fase", "tirada", "destino")),
)


def resolve_pliego_template_path(templates_dir: Path) -> Optional[Path]:
    for name in PLIEGO_TEMPLATE_FILENAMES:
        p = templates_dir / name
        if p.is_file():
            return p
    # Repo: docs/provided_docs (plantilla oficial si aún no está copiada a app/templates)
    repo_root = Path(__file__).resolve().parent.parent.parent.parent
    docs_provided = repo_root / "docs" / "provided_docs"
    for name in PLIEGO_TEMPLATE_FILENAMES:
        p = docs_provided / name
        if p.is_file():
            return p
    return None


def suggested_pliego_xlsx_filename(project_uuid: str) -> str:
    return f"GA-FO-01-(06-2025)-V02- Pliego de Condiciones - Arquitectura-{project_uuid}.xlsx"


def _strip_accents(s: str) -> str:
    nk = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nk if not unicodedata.combining(c))


def _norm_header(s: str) -> str:
    t = _strip_accents(str(s).strip().lower())
    t = re.sub(r"\s+", " ", t)
    return t


def _cell_display_value(ws: Worksheet, row: int, col: int) -> Optional[str]:
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for m in ws.merged_cells.ranges:
            if cell.coordinate in m:
                top = ws.cell(row=m.min_row, column=m.min_col)
                v = top.value
                return str(v).strip() if v is not None else None
        return None
    v = cell.value
    if v is None:
        return None
    return str(v).strip()


def _find_header_row(ws: Worksheet, max_row: int = 120) -> Optional[int]:
    best_row: Optional[int] = None
    best_score = 0
    limit = min(ws.max_row, max_row)
    for r in range(1, limit + 1):
        score = 0
        texts: list[str] = []
        for c in range(1, min(ws.max_column + 1, 40)):
            raw = _cell_display_value(ws, r, c)
            if raw:
                texts.append(_norm_header(raw))
        joined = " ".join(texts)
        if "descripcion" in joined:
            score += 3
        if any(k in joined for k in ("partida", "item", "n°", "nº")):
            score += 2
        if any(x in joined for x in ("cantidad", "cant ")):
            score += 2
        if any(x in joined for x in ("unidad", "und")):
            score += 1
        if "precio" in joined or "unit" in joined:
            score += 1
        if "subtotal" in joined or "importe" in joined:
            score += 1
        if score > best_score:
            best_score = score
            best_row = r
    if best_score >= 5:
        return best_row
    return None


def _map_columns(ws: Worksheet, header_row: int) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for c in range(1, min(ws.max_column + 1, 40)):
        raw = _cell_display_value(ws, header_row, c)
        if not raw:
            continue
        nh = _norm_header(raw)
        for field, keywords in HEADER_KEYWORDS:
            if field in col_map:
                continue
            for kw in keywords:
                if kw in nh or nh in kw:
                    col_map[field] = c
                    break
    return col_map


def _find_footer_row(ws: Worksheet, data_start: int, max_scan: int = 400) -> Optional[int]:
    end = min(ws.max_row + 1, data_start + max_scan)
    for r in range(data_start, end):
        for c in range(1, min(4, ws.max_column + 1)):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            u = str(v).strip().upper()
            if u.startswith("TOTAL") and "DESCRIPCI" not in u:
                return r
    return None


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def _clear_row_values(ws: Worksheet, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        if isinstance(cell, MergedCell):
            continue
        cell.value = None


def _sort_groups(groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(groups, key=lambda g: g.get("order", 0))


def _flatten_rows(groups: list[dict[str, Any]]) -> list[Tuple[str, Optional[dict[str, Any]], Optional[dict[str, Any]]]]:
    out: list[Tuple[str, Optional[dict[str, Any]], Optional[dict[str, Any]]]] = []
    for g in _sort_groups(groups):
        out.append(("section", g, None))
        for it in g.get("items", []):
            out.append(("item", g, it))
    return out


def _write_item(
    ws: Worksheet,
    row: int,
    col_map: dict[str, int],
    item: dict[str, Any],
) -> None:
    if "partida" in col_map:
        p = item.get("partida") or item.get("id")
        ws.cell(row=row, column=col_map["partida"], value=p if p is not None else None)
    if "descripcion" in col_map:
        ws.cell(row=row, column=col_map["descripcion"], value=item.get("descripcion"))
    if "capitulo" in col_map:
        ws.cell(row=row, column=col_map["capitulo"], value=item.get("capitulo"))
    if "unidad" in col_map:
        ws.cell(row=row, column=col_map["unidad"], value=item.get("unidad"))
    if "cantidad" in col_map:
        ws.cell(row=row, column=col_map["cantidad"], value=item.get("cantidad"))
    if "precio_unitario" in col_map:
        ws.cell(row=row, column=col_map["precio_unitario"], value=item.get("precio_unitario"))
    if "subtotal" in col_map:
        ws.cell(row=row, column=col_map["subtotal"], value=item.get("subtotal"))
    if "notas" in col_map:
        ws.cell(row=row, column=col_map["notas"], value=item.get("notas"))


def _write_section(
    ws: Worksheet,
    row: int,
    col_map: dict[str, int],
    group: dict[str, Any],
) -> None:
    label = f"{group.get('title', '')} ({group.get('kind', '')})"
    if "grupo" in col_map:
        ws.cell(row=row, column=col_map["grupo"], value=label)
        return
    if "capitulo" in col_map:
        ws.cell(row=row, column=col_map["capitulo"], value=label)
        return
    if "descripcion" in col_map:
        ws.cell(row=row, column=col_map["descripcion"], value=label)
        return
    ws.cell(row=row, column=1, value=label)


def _lookup_ga_fo_01_item_state(cell_a: Any, item_states: dict[str, Any]) -> Optional[dict[str, Any]]:
    if cell_a is None:
        return None
    if isinstance(cell_a, float):
        s = str(cell_a).rstrip("0").rstrip(".") if cell_a == int(cell_a) else str(cell_a)
    else:
        s = str(cell_a).strip()
    if not s:
        return None
    candidates = [s, s.rstrip("."), s if s.endswith(".") else f"{s}."]
    for key in candidates:
        if key in item_states:
            row = item_states[key]
            return row if isinstance(row, dict) else None
    return None


def _export_estado_label(raw: Any) -> str:
    if raw is None:
        return ""
    code = str(raw).strip().upper()
    return GA_FO_01_ESTADO_LABELS.get(code, str(raw).strip())


def fill_resumen_pliego_ga_fo_01(wb: Workbook, item_states: dict[str, Any]) -> bool:
    """
    Escribe estado y observaciones en la hoja RESUMEN del Excel oficial,
    emparejando la columna A (N.º) con las claves del JSON guardado en la app.
    """
    if not item_states:
        return False
    if RESUMEN_SHEET_NAME not in wb.sheetnames:
        return False
    ws = wb[RESUMEN_SHEET_NAME]
    filled = 0
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1)
        if isinstance(cell_a, MergedCell):
            continue
        st = _lookup_ga_fo_01_item_state(cell_a.value, item_states)
        if st is None:
            continue
        estado = _export_estado_label(st.get("estado"))
        obs_parts: list[str] = []
        notas = st.get("notas")
        if isinstance(notas, str) and notas.strip():
            obs_parts.append(notas.strip())
        fn = st.get("file_name")
        if isinstance(fn, str) and fn.strip():
            obs_parts.append(f"Archivo adjunto: {fn.strip()}")
        observaciones = "\n".join(obs_parts) if obs_parts else ""

        c_est = ws.cell(row=row, column=RESUMEN_COL_ESTADO)
        if not isinstance(c_est, MergedCell):
            c_est.value = estado
        c_obs = ws.cell(row=row, column=RESUMEN_COL_OBSERVACIONES)
        if not isinstance(c_obs, MergedCell):
            c_obs.value = observaciones
        filled += 1
    return filled > 0


def fill_pliego_workbook(wb: Workbook, payload: dict[str, Any]) -> bool:
    groups = payload.get("groups") or []
    if not isinstance(groups, list):
        return False

    ws = wb.active
    header_row = _find_header_row(ws)
    if header_row is None:
        return False

    col_map = _map_columns(ws, header_row)
    if "descripcion" not in col_map and "partida" not in col_map:
        return False

    max_col = max(ws.max_column, max(col_map.values(), default=1))
    data_start = header_row + 1
    flat = _flatten_rows(groups)
    needed = len(flat)
    if needed == 0:
        return True

    footer_row = _find_footer_row(ws, data_start)
    template_style_row = data_start
    if footer_row is not None:
        available = footer_row - data_start
        if needed > available:
            ins = needed - available
            ws.insert_rows(footer_row, ins)
            footer_row += ins
            for i in range(ins):
                r = footer_row - ins + i
                _copy_row_style(ws, template_style_row, r, max_col)
        last_data = data_start + needed - 1
        for r in range(last_data + 1, footer_row):
            _clear_row_values(ws, r, max_col)
    for i, (kind, g, it) in enumerate(flat):
        r = data_start + i
        while r > ws.max_row:
            ws.append([None] * max_col)
        _copy_row_style(ws, template_style_row, r, max_col)
        _clear_row_values(ws, r, max_col)
        assert g is not None
        if kind == "section":
            _write_section(ws, r, col_map, g)
        elif it is not None:
            _write_item(ws, r, col_map, it)

    if footer_row is not None:
        for r in range(data_start + needed, footer_row):
            _clear_row_values(ws, r, max_col)

    return True


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
