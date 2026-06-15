from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import Any, Optional
from uuid import UUID

from fpdf import FPDF
from fpdf.enums import XPos
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.models.plan_delivery_request import PlanDeliveryRequest
from app.models.project_file import ProjectFile
from app.models.user import User
from app.services.pliego_template_fill import (
    fill_pliego_workbook,
    fill_resumen_pliego_ga_fo_01,
    resolve_pliego_template_path,
    suggested_pliego_xlsx_filename,
    workbook_to_bytes,
)
from app.services.plan_delivery_service import PlanDeliveryService
from app.services.project_service import ProjectService

settings = get_settings()


def _control_days_display(row: PlanDeliveryRequest) -> str:
    if row.days_count is not None:
        return str(row.days_count)
    if row.request_date is not None and row.delivery_date is not None:
        return str((row.delivery_date - row.request_date).days)
    return ""


class ExportService:
    def __init__(self, session: AsyncSession, workspace_id: UUID) -> None:
        self._session = session
        self._workspace_id = workspace_id
        self._project_service = ProjectService(session, workspace_id)

    async def _load_payload(self, user: User, project_uuid: UUID) -> dict:
        payload, _ = await self._project_service.get_architecture(user, project_uuid)
        return payload

    def build_pliego_xlsx(self, payload: dict) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pliego"
        headers = ["Grupo", "Tipo", "Ítem", "Descripción", "Unidad", "Cant.", "P. Unit.", "Subtotal", "Notas"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
        row_idx = 2
        for g in payload.get("groups", []):
            gtitle = g.get("title", "")
            gkind = g.get("kind", "")
            for it in g.get("items", []):
                ws.cell(row=row_idx, column=1, value=gtitle)
                ws.cell(row=row_idx, column=2, value=gkind)
                ws.cell(row=row_idx, column=3, value=str(it.get("partida", "") or it.get("id", "")))
                ws.cell(row=row_idx, column=4, value=it.get("descripcion", ""))
                ws.cell(row=row_idx, column=5, value=it.get("unidad", ""))
                ws.cell(row=row_idx, column=6, value=it.get("cantidad"))
                ws.cell(row=row_idx, column=7, value=it.get("precio_unitario"))
                ws.cell(row=row_idx, column=8, value=it.get("subtotal"))
                ws.cell(row=row_idx, column=9, value=it.get("notas", ""))
                row_idx += 1
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def build_control_planos_xlsx(self, project_name: str, rows: list[PlanDeliveryRequest]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Control Planos"
        headers = [
            "No.",
            "Fecha de Solicitud",
            "Proyecto",
            "No. Solicitud",
            "Descripción",
            "Fecha de Entrega",
            "Cantidad día(s)",
            "Estado",
        ]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True)
        for idx, row in enumerate(rows, start=1):
            r = idx + 1
            ws.cell(r, 1, value=idx)
            ws.cell(r, 2, value=row.request_date.isoformat() if row.request_date else "")
            ws.cell(r, 3, value=project_name)
            ws.cell(r, 4, value=row.request_number)
            ws.cell(r, 5, value=row.description)
            ws.cell(r, 6, value=row.delivery_date.isoformat() if row.delivery_date else "")
            ws.cell(r, 7, value=_control_days_display(row))
            ws.cell(r, 8, value=row.status)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def build_pdf(self, title: str, payload: dict) -> bytes:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(0, 10, title, ln=True)
        pdf.set_font("Helvetica", size=10)
        for g in payload.get("groups", []):
            pdf.set_font("Helvetica", "B", 11)
            pdf.multi_cell(0, 7, f"{g.get('title', '')} ({g.get('kind', '')})", new_x=XPos.LMARGIN)
            pdf.set_font("Helvetica", size=9)
            for it in g.get("items", []):
                line = (
                    f"{it.get('partida', '')} | {it.get('descripcion', '')} | "
                    f"{it.get('unidad', '')} | {it.get('cantidad', '')} | "
                    f"{it.get('precio_unitario', '')} | {it.get('subtotal', '')}"
                )
                pdf.multi_cell(0, 6, line, new_x=XPos.LMARGIN)
            pdf.ln(2)
        out = pdf.output()
        if isinstance(out, (bytes, bytearray)):
            return bytes(out)
        return str(out).encode("latin-1", errors="replace")

    def build_documentary_report_pdf(
        self,
        *,
        project_name: str,
        project_code: Optional[str],
        location_text: Optional[str],
        client_name: Optional[str],
        criteria: list[dict[str, Any]],
        files: list[ProjectFile],
    ) -> bytes:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 14)
        pdf.multi_cell(0, 8, "Informe documental - checklist y archivos", new_x=XPos.LMARGIN)
        pdf.set_font("Helvetica", size=10)
        pdf.ln(2)
        meta_lines = [
            f"Proyecto: {project_name}",
            f"Código: {project_code or '-'}",
            f"Cliente: {client_name or '-'}",
            f"Ubicación: {location_text or '-'}",
        ]
        for line in meta_lines:
            pdf.multi_cell(0, 6, line, new_x=XPos.LMARGIN)

        required = [c for c in criteria if isinstance(c, dict) and c.get("required")]
        done_req = sum(1 for c in required if c.get("done"))
        total_req = len(required)
        pct = int(round(100 * done_req / total_req)) if total_req else 0
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 11)
        pdf.multi_cell(0, 7, "Cumplimiento checklist (ítems requeridos)", new_x=XPos.LMARGIN)
        pdf.set_font("Helvetica", size=9)
        pdf.multi_cell(0, 6, f"{done_req} de {total_req} completados ({pct}%).", new_x=XPos.LMARGIN)
        for item in criteria:
            if not isinstance(item, dict):
                continue
            lab = str(item.get("label", "")).strip() or "Ítem"
            req = "Sí" if item.get("required") else "No"
            ok = "Listo" if item.get("done") else "Pendiente"
            pdf.multi_cell(0, 5, f"- [{ok}] {lab} (obligatorio: {req})", new_x=XPos.LMARGIN)

        names_norm = [f.original_name.strip().lower() for f in files]
        dup_counts = Counter(names_norm)
        duplicates = [n for n, k in dup_counts.items() if k > 1 and n]

        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 11)
        pdf.multi_cell(0, 7, f"Archivos cargados ({len(files)})", new_x=XPos.LMARGIN)
        pdf.set_font("Helvetica", size=9)
        if duplicates:
            pdf.set_font("Helvetica", "B", 9)
            pdf.multi_cell(
                0,
                5,
                "Posibles duplicados por nombre (misma denominación): "
                + ", ".join(duplicates[:25])
                + ("…" if len(duplicates) > 25 else ""),
                new_x=XPos.LMARGIN,
            )
            pdf.set_font("Helvetica", size=9)
        if not files:
            pdf.multi_cell(0, 6, "No hay archivos registrados.", new_x=XPos.LMARGIN)
        else:
            for f in files:
                disc = f.discipline or "-"
                pdf.multi_cell(
                    0,
                    5,
                    f"- {f.original_name} | disciplina: {disc} | estado ingestión: {f.ingest_status}",
                    new_x=XPos.LMARGIN,
                )

        out = pdf.output()
        if isinstance(out, (bytes, bytearray)):
            return bytes(out)
        return str(out).encode("latin-1", errors="replace")

    async def export_documentary_pdf(self, user: User, project_uuid: UUID) -> bytes:
        project = await self._project_service.get_project(user, project_uuid)
        q = (
            select(ProjectFile)
            .where(ProjectFile.project_id == project.id)
            .order_by(ProjectFile.created_at.asc())
        )
        rows = list((await self._session.execute(q)).scalars().all())
        crit = project.project_bootstrap_criteria or []
        if not isinstance(crit, list):
            crit = []
        return self.build_documentary_report_pdf(
            project_name=project.name,
            project_code=project.project_code,
            location_text=project.location_text,
            client_name=project.client_name,
            criteria=crit,
            files=rows,
        )

    async def export_pliego_xlsx(self, user: User, project_uuid: UUID) -> tuple[bytes, str]:
        project = await self._project_service.get_project(user, project_uuid)
        spec = project.specifications_document or {}
        ga = spec.get("ga_fo_01_arquitectura")
        item_states: dict[str, Any] = {}
        if isinstance(ga, dict):
            raw = ga.get("item_states")
            if isinstance(raw, dict):
                item_states = raw

        payload = await self._load_payload(user, project_uuid)
        tpl = resolve_pliego_template_path(Path(settings.templates_dir))
        if tpl is not None:
            wb = load_workbook(tpl)
            if item_states and fill_resumen_pliego_ga_fo_01(wb, item_states):
                return workbook_to_bytes(wb), suggested_pliego_xlsx_filename(str(project_uuid))
            if fill_pliego_workbook(wb, payload):
                return workbook_to_bytes(wb), suggested_pliego_xlsx_filename(str(project_uuid))
        return self.build_pliego_xlsx(payload), f"pliego-{project_uuid}.xlsx"

    async def export_control_xlsx(self, user: User, project_uuid: UUID) -> bytes:
        pds = PlanDeliveryService(self._session, self._workspace_id)
        project_name, rows = await pds.list_models_for_export(user, project_uuid)
        return self.build_control_planos_xlsx(project_name, rows)

    async def export_pliego_pdf(self, user: User, project_uuid: UUID) -> bytes:
        payload = await self._load_payload(user, project_uuid)
        return self.build_pdf("Pliego de Condiciones - Arquitectura", payload)

    async def export_control_pdf(self, user: User, project_uuid: UUID) -> bytes:
        pds = PlanDeliveryService(self._session, self._workspace_id)
        project_name, rows = await pds.list_models_for_export(user, project_uuid)
        return self.build_control_planos_pdf(project_name, rows)

    def build_control_planos_pdf(self, project_name: str, rows: list[PlanDeliveryRequest]) -> bytes:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 12)
        pdf.multi_cell(0, 8, f"Control Entrega de Planos - {project_name}", new_x=XPos.LMARGIN)
        pdf.ln(2)
        pdf.set_font("Helvetica", size=8)
        if not rows:
            pdf.cell(0, 6, "Sin solicitudes registradas.", ln=True)
        else:
            for idx, row in enumerate(rows, start=1):
                pdf.set_font("Helvetica", "B", 9)
                pdf.multi_cell(
                    0,
                    5,
                    f"{idx}. {row.request_number} - {row.status}",
                    new_x=XPos.LMARGIN,
                )
                pdf.set_font("Helvetica", size=8)
                line = (
                    f"Solicitud: {row.request_date.isoformat() if row.request_date else '-'} | "
                    f"Entrega: {row.delivery_date.isoformat() if row.delivery_date else '-'} | "
                    f"Días: {_control_days_display(row) or '-'}"
                )
                pdf.multi_cell(0, 4, line, new_x=XPos.LMARGIN)
                desc = (row.description or "").strip() or "-"
                pdf.multi_cell(0, 4, desc, new_x=XPos.LMARGIN)
                pdf.ln(1)
        out = pdf.output()
        if isinstance(out, (bytes, bytearray)):
            return bytes(out)
        return str(out).encode("latin-1", errors="replace")
