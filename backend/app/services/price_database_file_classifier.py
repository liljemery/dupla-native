"""Clasificación IA: materiales | mano_obra | subcontratos (base de precios por proyecto)."""

from __future__ import annotations

import csv
import json
import logging
from io import StringIO
from pathlib import Path
from typing import Optional

from openai import AsyncOpenAI
from openpyxl import load_workbook
from pypdf import PdfReader

from app.config import Settings

logger = logging.getLogger(__name__)

PRICE_CATEGORIES = frozenset({"materiales", "mano_obra", "subcontratos"})

_SYSTEM = """Eres experto en construcción y presupuestos en España/Latam.
Clasifica el documento en UNA de estas categorías de base de precios:
- materiales: listados o tablas de precios unitarios de insumos, materiales, herramientas, equipos.
- mano_obra: tabuladores salariales, costes hora hombre, convenios, extras, prestaciones.
- subcontratos: cotizaciones de empresas externas, especialistas, servicios subcontratados.

Responde SOLO JSON válido con claves:
- price_category: uno exactamente de: materiales, mano_obra, subcontratos
- confidence: número 0..1
- reason_short: una frase corta en español
Si no puedes decidir, usa la categoría más probable y confidence baja."""


def extract_pdf_snippet(file_path: Path, max_chars: int = 6000) -> str:
    if file_path.suffix.lower() != ".pdf" or not file_path.is_file():
        return ""
    try:
        reader = PdfReader(str(file_path))
        parts: list[str] = []
        total = 0
        for page in reader.pages[:6]:
            t = (page.extract_text() or "").strip()
            if not t:
                continue
            parts.append(t)
            total += len(t)
            if total >= max_chars:
                break
        return "\n\n".join(parts)[:max_chars]
    except Exception as exc:
        logger.debug("price_db PDF snippet: %s", exc)
        return ""


def extract_csv_snippet(file_path: Path, max_chars: int = 6000) -> str:
    if file_path.suffix.lower() != ".csv" or not file_path.is_file():
        return ""
    try:
        raw = file_path.read_text(encoding="utf-8", errors="replace")
        return raw[:max_chars]
    except Exception as exc:
        logger.debug("price_db CSV snippet: %s", exc)
        return ""


def extract_xlsx_snippet(file_path: Path, max_chars: int = 6000) -> str:
    ext = file_path.suffix.lower()
    if ext not in (".xlsx", ".xls") or not file_path.is_file():
        return ""
    if ext == ".xls":
        return ""
    try:
        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows: list[str] = []
            n = 0
            for row in ws.iter_rows(max_row=45, max_col=20, values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                line = " | ".join(cells[:12])
                if line.strip():
                    rows.append(line)
                n += 1
                joined = "\n".join(rows)
                if len(joined) >= max_chars:
                    return joined[:max_chars]
            return "\n".join(rows)[:max_chars]
        finally:
            wb.close()
    except Exception as exc:
        logger.debug("price_db xlsx snippet: %s", exc)
        return ""


def build_text_preview(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_pdf_snippet(path)
    if ext == ".csv":
        return extract_csv_snippet(path)
    if ext in (".xlsx", ".xls"):
        return extract_xlsx_snippet(path)
    return ""


async def classify_price_database_file(
    settings: Settings,
    *,
    original_name: str,
    mime: Optional[str],
    file_size: int,
    text_preview: str,
) -> tuple[Optional[str], float, str]:
    if not (settings.openai_api_key or "").strip():
        return None, 0.0, "OPENAI_API_KEY no configurada"

    client = AsyncOpenAI(api_key=settings.openai_api_key)
    preview = (text_preview or "").strip()
    if len(preview) > 12000:
        preview = preview[:12000] + "…"

    user_content = json.dumps(
        {
            "original_name": original_name,
            "mime": mime,
            "file_size": file_size,
            "sheet_or_text_preview": preview,
        },
        ensure_ascii=False,
    )

    try:
        resp = await client.chat.completions.create(
            model=settings.openai_model,
            temperature=0.1,
            messages=[
                {"role": "system", "content": _SYSTEM},
                {"role": "user", "content": user_content},
            ],
            response_format={"type": "json_object"},
        )
        raw = (resp.choices[0].message.content or "").strip()
        data = json.loads(raw)
    except Exception as exc:
        logger.warning("price_db classify OpenAI: %s", exc)
        return None, 0.0, str(exc)[:500]

    cat = data.get("price_category")
    conf = data.get("confidence")
    reason = data.get("reason_short")
    if not isinstance(cat, str) or cat.strip() not in PRICE_CATEGORIES:
        return None, 0.0, "Categoría inválida en respuesta IA"
    try:
        c = float(conf) if conf is not None else 0.0
    except (TypeError, ValueError):
        c = 0.0
    r = reason if isinstance(reason, str) else ""
    return cat.strip(), max(0.0, min(1.0, c)), r[:400]
