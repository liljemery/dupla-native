"""Tests unitarios del relleno GA-FO-01 (sin base de datos)."""

import io
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from app.services.pliego_template_fill import (
    fill_pliego_workbook,
    resolve_pliego_template_path,
    suggested_pliego_xlsx_filename,
    workbook_to_bytes,
)


def _minimal_ga_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pliego"
    # Cabecera tipo GA (fila 8)
    ws.cell(row=8, column=1, value="Partida")
    ws.cell(row=8, column=2, value="Descripción")
    ws.cell(row=8, column=3, value="Unidad")
    ws.cell(row=8, column=4, value="Cantidad")
    ws.cell(row=8, column=5, value="Precio unitario")
    ws.cell(row=8, column=6, value="Subtotal")
    # Fila de datos plantilla + TOTAL
    ws.cell(row=9, column=1, value="")
    ws.cell(row=20, column=1, value="TOTAL")
    return wb


def test_fill_pliego_workbook_writes_rows():
    wb = _minimal_ga_workbook()
    gid = str(uuid4())
    iid = str(uuid4())
    payload = {
        "groups": [
            {
                "id": gid,
                "kind": "fase",
                "title": "Fase I",
                "order": 0,
                "items": [
                    {
                        "id": iid,
                        "descripcion": "Muro",
                        "partida": "1.01",
                        "unidad": "m2",
                        "cantidad": 10,
                        "precio_unitario": 5,
                        "subtotal": 50,
                        "notas": None,
                        "capitulo": None,
                    }
                ],
            }
        ]
    }
    assert fill_pliego_workbook(wb, payload) is True
    ws = wb.active
    # section + item => filas 9 y 10
    assert "Fase I" in str(ws.cell(row=9, column=2).value) or "Fase I" in str(ws.cell(row=9, column=1).value)
    assert ws.cell(row=10, column=1).value == "1.01"
    assert ws.cell(row=10, column=2).value == "Muro"


def test_suggested_filename_contains_uuid():
    u = str(uuid4())
    assert u in suggested_pliego_xlsx_filename(u)
    assert "GA-FO-01" in suggested_pliego_xlsx_filename(u)


def test_resolve_pliego_template_missing(tmp_path):
    assert resolve_pliego_template_path(tmp_path, search_repo_fallback=False) is None


def test_roundtrip_bytes():
    wb = _minimal_ga_workbook()
    payload = {
        "groups": [
            {
                "id": str(uuid4()),
                "kind": "tirada",
                "title": "G",
                "order": 0,
                "items": [
                    {
                        "id": str(uuid4()),
                        "descripcion": "X",
                        "partida": "1",
                        "unidad": "ud",
                        "cantidad": 1,
                        "precio_unitario": 1,
                        "subtotal": 1,
                        "notas": None,
                        "capitulo": None,
                    }
                ],
            }
        ]
    }
    fill_pliego_workbook(wb, payload)
    raw = workbook_to_bytes(wb)
    wb2 = load_workbook(io.BytesIO(raw))
    assert fill_pliego_workbook(wb2, payload) is True
